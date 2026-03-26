from fastapi import FastAPI, Depends, HTTPException, status, Response, WebSocket, WebSocketDisconnect, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import func, text
from contextlib import asynccontextmanager
from datetime import timedelta, datetime
import cv2
import json
import asyncio
import queue
import signal
import logging
import os
from typing import List
from pydantic import BaseModel
from . import models, schemas, database, auth, vision_engine
from .database import engine, SessionLocal, get_db

# ─── Logging structuré ────────────────────────────────────────────────────────
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO").upper(),
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%S",
)
logger = logging.getLogger("ciment")

models.Base.metadata.create_all(bind=engine)

# ── Auto-migration : add columns added after initial schema creation ──────────
with engine.connect() as _mig_conn:
    for _mig_sql in [
        "ALTER TABLE detection_logs ADD COLUMN is_resolved BOOLEAN DEFAULT 0",
    ]:
        try:
            _mig_conn.execute(text(_mig_sql))
            _mig_conn.commit()
        except Exception:
            pass  # column already exists


# ─── WebSocket Manager (events: COUNT_EVENT, etc.) ───────────────────────────
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast(self, message: str):
        dead = []
        for connection in self.active_connections:
            try:
                await connection.send_text(message)
            except Exception:
                dead.append(connection)
        for c in dead:
            self.disconnect(c)

manager = ConnectionManager()

# Reference to the main asyncio event loop (set at startup)
_main_loop: asyncio.AbstractEventLoop = None


# ─── Vision Event Callback (called from vision thread) ───────────────────────
def handle_vision_event(event_data):
    db = SessionLocal()
    try:
        db_log = models.DetectionLog(
            session_id=event_data["session_id"],
            status=event_data["status"],
            identifier=event_data["identifier"],
            detection_score=event_data["detection_score"],
            logo_score=event_data["logo_score"],
            color_score=event_data["color_score"],
            interval=event_data["interval"],
            capture_url=event_data["capture_url"]
        )
        db.add(db_log)

        session = db.query(models.Session).filter(models.Session.id == event_data["session_id"]).first()
        # Guard: ignore race-condition events that arrive after the session was stopped
        if session and session.status != "active":
            logger.warning(
                "handle_vision_event: session %s is already '%s', ignoring late event",
                event_data["session_id"], session.status,
            )
            return
        if session:
            if event_data["status"] == "conforme":
                session.total_count += 1
            else:
                session.rejected_count += 1

        db.commit()

        message = {
            "type": "COUNT_EVENT",
            "data": {
                "id": db_log.id,
                "status": db_log.status,
                "identifier": db_log.identifier,
                "timestamp": db_log.timestamp.isoformat(),
                "detection_score": db_log.detection_score,
                "logo_score": db_log.logo_score,
                "color_score": db_log.color_score,
                "capture_url": db_log.capture_url,
                "session_stats": {
                    "total": session.total_count if session else 0,
                    "rejected": session.rejected_count if session else 0
                }
            }
        }

        # Thread-safe broadcast to the asyncio event loop
        if _main_loop and _main_loop.is_running():
            asyncio.run_coroutine_threadsafe(
                manager.broadcast(json.dumps(message)),
                _main_loop
            )

    except Exception as e:
        logger.error("Vision event handling error: %s", e)
    finally:
        db.close()


# ─── Lifespan: replaces deprecated @app.on_event ─────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    global _main_loop
    _main_loop = asyncio.get_running_loop()

    # Load camera config from DB and apply to engine
    db = SessionLocal()
    try:
        settings = db.query(models.SystemSetting).filter(
            models.SystemSetting.key.in_([
                "camera_source_type", "camera_url", "camera_resolution", "camera_fps", "camera_brightness", "camera_contrast", "camera_autofocus",
                "detection_model_path", "detection_threshold", "detection_nms_iou", "detection_max_det", "detection_imgsz", "tracking_persistence",
                "virtual_line_y_percent", "virtual_line_span_percent", "virtual_line_direction"
            ])
        ).all()
        config = {s.key: s.value for s in settings}
        source_type = config.get("camera_source_type", "webcam")
        url = config.get("camera_url", "0")

        v_engine = vision_engine.get_vision_engine()
        if source_type == "webcam":
            v_engine.video_source = int(url) if url.isdigit() else 0
        else:
            v_engine.video_source = url

        v_engine.apply_camera_settings(
            resolution=config.get("camera_resolution", "720p"),
            fps=int(config.get("camera_fps", "30")),
            brightness=int(config.get("camera_brightness", "50")),
            contrast=int(config.get("camera_contrast", "65")),
            autofocus=config.get("camera_autofocus", "true").lower() == "true",
        )

        v_engine.apply_model_config(
            model_path=config.get("detection_model_path", "models/best_V5.pt"),
            confidence_threshold=float(config.get("detection_threshold", "0.7")),
            nms_iou_threshold=float(config.get("detection_nms_iou", "0.45")),
            max_detections=int(config.get("detection_max_det", "100")),
            inference_size=int(config.get("detection_imgsz", "1280")),
            tracking_persistence=config.get("tracking_persistence", "true").lower() == "true",
        )
        v_engine.apply_virtual_line_config(
            position_percent=int(config.get("virtual_line_y_percent", "60")),
            line_span_percent=int(config.get("virtual_line_span_percent", "80")),
            direction=config.get("virtual_line_direction", "left-right"),
        )

        v_engine.set_on_count_callback(handle_vision_event)
        v_engine.start()
    finally:
        db.close()

    # ── DB Migration: add missing columns (SQLite ALTER TABLE) ───────────────
    from sqlalchemy import text as _sa_text
    for _stmt in [
        "ALTER TABLE alert_history ADD COLUMN alert_type TEXT NOT NULL DEFAULT 'info'",
        "ALTER TABLE alert_history ADD COLUMN title TEXT",
        "ALTER TABLE users ADD COLUMN last_login DATETIME",
        "ALTER TABLE users ADD COLUMN login_count INTEGER DEFAULT 0",
    ]:
        try:
            with engine.connect() as _conn:
                _conn.execute(_sa_text(_stmt))
                _conn.commit()
        except Exception:
            pass  # Column already exists

    # ── Load template & colour config and apply to vision engine ─────────────
    _db_tmpl = SessionLocal()
    try:
        import json as _json_tmpl
        _ts = {s.key: s.value for s in _db_tmpl.query(models.SystemSetting).filter(
            models.SystemSetting.key.in_([
                "template_active_file", "template_threshold",
                "template_colors", "template_color_threshold",
            ])
        ).all()}
        _tmpl_file = _ts.get("template_active_file", "")
        _tmpl_path = f"backend/static/templates/{_tmpl_file}" if _tmpl_file else None
        _threshold  = float(_ts.get("template_threshold", "0.65"))
        _color_thr  = float(_ts.get("template_color_threshold", "0.25"))
        _color_refs = _json_tmpl.loads(_ts.get("template_colors", "[]"))
        v_engine.get_vision_engine().apply_template_config(_tmpl_path, _threshold, _color_refs, _color_thr)
    except Exception as _e:
        logger.warning("Template config load failed: %s", _e)
    finally:
        _db_tmpl.close()

    # ── Seed cameras table from legacy SystemSettings (backward compat) ──────
    _db_seed = SessionLocal()
    try:
        if _db_seed.query(models.Camera).count() == 0:
            _s = {s.key: s.value for s in _db_seed.query(models.SystemSetting).filter(
                models.SystemSetting.key.in_([
                    "camera_source_type", "camera_url", "camera_resolution", "camera_fps"
                ])
            ).all()}
            _db_seed.add(models.Camera(
                name="Caméra Principale",
                source_type=_s.get("camera_source_type", "webcam"),
                url=_s.get("camera_url", "0"),
                resolution=_s.get("camera_resolution", "720p"),
                fps=int(_s.get("camera_fps", "30")),
                is_active=True,
            ))
            _db_seed.commit()
    except Exception as _e:
        logger.warning("Camera seed failed: %s", _e)
    finally:
        _db_seed.close()

    # ── Seed default roles ──────────────────────────────────────────────────
    _db_roles = SessionLocal()
    try:
        import json as _json_roles
        for _rd in _DEFAULT_ROLES_SEED:
            if not _db_roles.query(models.Role).filter(models.Role.name == _rd["name"]).first():
                _db_roles.add(models.Role(
                    name=_rd["name"], label=_rd["label"],
                    description=_rd["description"],
                    permissions=_json_roles.dumps(_rd["permissions"]),
                    is_builtin=_rd["is_builtin"],
                ))
        _db_roles.commit()
    except Exception as _re:
        logger.warning("Role seeding failed: %s", _re)
    finally:
        _db_roles.close()

    # ── Seed default admin user (create if absent, never overwrite existing) ──
    _db_admin = SessionLocal()
    try:
        _existing_admin = _db_admin.query(models.User).filter(models.User.username == "admin").first()
        if not _existing_admin:
            _db_admin.add(models.User(
                username="admin",
                hashed_password=auth.get_password_hash("admin1234"),
                full_name="Administrateur",
                role="admin",
                is_active=True,
            ))
            _db_admin.commit()
            logger.info("Compte admin par défaut créé  →  admin / admin1234")
        else:
            # Ensure the admin role is correct (never downgrade an existing account)
            if _existing_admin.role != "admin":
                _existing_admin.role = "admin"
                _db_admin.commit()
    except Exception as _ae:
        logger.warning("Admin seed failed: %s", _ae)
    finally:
        _db_admin.close()

    # ── Start scheduled export background task ────────────────────────────
    global _schedule_task
    _schedule_task = asyncio.create_task(_scheduler_loop())

    yield  # ── Application is running ──

    # Shutdown: stop the vision engine cleanly + cancel scheduler
    logger.info("Shutdown signal reçu, arrêt du moteur de vision...")
    v_engine = vision_engine.get_vision_engine()
    v_engine.stop()
    if _schedule_task and not _schedule_task.done():
        _schedule_task.cancel()
    _main_loop = None
    logger.info("Shutdown complet.")


app = FastAPI(title="Cement Bag Counter API", version="1.0.0", lifespan=lifespan)

app.mount("/static", StaticFiles(directory="backend/static"), name="static")

_allowed_origins_env = os.getenv("ALLOWED_ORIGINS", "")
_cors_origins = [o.strip() for o in _allowed_origins_env.split(",") if o.strip()]
if not _cors_origins:
    _cors_origins = ["*"]
    logger.warning(
        "ALLOWED_ORIGINS non défini — CORS accepte toutes les origines (*). "
        "Définissez ALLOWED_ORIGINS dans .env pour la production."
    )

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Root ─────────────────────────────────────────────────────────────────────
@app.get("/")
async def root():
    return {"message": "Welcome to Cement Bag Counter API"}


# ─── Health check (utilisé par Docker healthcheck) ────────────────────────────
from sqlalchemy import text as _sa_text_health

@app.get("/api/health")
async def health_check(db: Session = Depends(get_db)):
    try:
        db.execute(_sa_text_health("SELECT 1"))
    except Exception:
        raise HTTPException(status_code=503, detail="Database unavailable")
    return {"status": "ok", "version": "1.0.0"}


# ─── Video Feed: MJPEG fallback (kept for backwards compat) ──────────────────
def gen_frames():
    v_engine = vision_engine.get_vision_engine()
    while True:
        frame = v_engine.get_video_frame()
        if frame is not None:
            (flag, encoded) = cv2.imencode(".jpg", frame)
            if not flag:
                continue
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + bytearray(encoded) + b'\r\n')
        else:
            import time
            time.sleep(0.1)


@app.get("/api/vision/video_feed")
async def video_feed():
    from fastapi.responses import StreamingResponse
    return StreamingResponse(gen_frames(), media_type="multipart/x-mixed-replace; boundary=frame")


# ─── WebSocket Video Stream (/ws/video) ──────────────────────────────────────
@app.websocket("/ws/video")
async def websocket_video(websocket: WebSocket):
    await websocket.accept()

    v_engine = vision_engine.get_vision_engine()
    frame_queue = queue.Queue(maxsize=3)
    v_engine.add_video_subscriber(frame_queue)

    try:
        while True:
            try:
                # Wait for a frame from the vision engine (with timeout to check WS alive)
                frame_b64 = await asyncio.get_event_loop().run_in_executor(
                    None, lambda: frame_queue.get(timeout=1.0)
                )
                await websocket.send_text(frame_b64)
            except queue.Empty:
                # No frame available — send a ping to keep alive and detect disconnects
                try:
                    await websocket.send_text("")
                except Exception:
                    break
            except WebSocketDisconnect:
                break
            except Exception:
                break
    finally:
        v_engine.remove_video_subscriber(frame_queue)


# ─── WebSocket Events (/ws) ──────────────────────────────────────────────────
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)


def _recompute_session_counters(db: Session, session_id: str):
    session = db.query(models.Session).filter(models.Session.id == session_id).first()
    if not session:
        return
    total_conforme = db.query(models.DetectionLog).filter(
        models.DetectionLog.session_id == session_id,
        models.DetectionLog.status == "conforme"
    ).count()
    total_rejete = db.query(models.DetectionLog).filter(
        models.DetectionLog.session_id == session_id,
        models.DetectionLog.status == "rejete"
    ).count()
    session.total_count = total_conforme
    session.rejected_count = total_rejete



# ─── Production Reports ──────────────────────────────────────────────────────
@app.get("/api/reports/production")
async def get_production_report(period: str = "week", db: Session = Depends(get_db)):
    """Production report for day / week / month with real vs previous period comparison."""
    import math as _math
    from datetime import datetime as _dt, timedelta as _td

    PERIOD_HOURS = {"day": 24, "week": 168, "month": 720}
    hours = PERIOD_HOURS.get(period, 168)
    now = _dt.utcnow()
    start = now - _td(hours=hours)
    prev_start = start - _td(hours=hours)
    planned_seconds = hours * 3600
    TARGET_RATE = 1100  # bags/h

    def _to_dt(ts):
        if isinstance(ts, str):
            try:
                return _dt.fromisoformat(ts)
            except Exception:
                return None
        return ts

    # ── Current period ──────────────────────────────────────────────────────
    logs = (
        db.query(models.DetectionLog)
        .filter(models.DetectionLog.timestamp >= start)
        .order_by(models.DetectionLog.timestamp.asc())
        .all()
    )
    conforming = [l for l in logs if l.status == "conforme"]
    rejected   = [l for l in logs if l.status == "rejete"]
    total_bags     = len(conforming)
    rejected_count = len(rejected)
    total_inspected = total_bags + rejected_count

    detection_rate = round((total_bags / total_inspected * 100) if total_inspected > 0 else 100.0, 1)

    all_ts = sorted(ts for l in conforming if (ts := _to_dt(l.timestamp)) is not None)
    intervals: list[float] = []
    for i in range(1, len(all_ts)):
        d = (all_ts[i] - all_ts[i - 1]).total_seconds()
        if 0 < d < 120:
            intervals.append(d)
    avg_interval = round(sum(intervals) / len(intervals), 2) if intervals else 0.0

    # ── Previous period ──────────────────────────────────────────────────────
    prev_logs = db.query(models.DetectionLog).filter(
        models.DetectionLog.timestamp >= prev_start,
        models.DetectionLog.timestamp < start,
    ).all()
    prev_conforming = [l for l in prev_logs if l.status == "conforme"]
    prev_total      = len(prev_conforming)
    prev_rejected   = len([l for l in prev_logs if l.status == "rejete"])
    prev_inspected  = prev_total + prev_rejected
    prev_detection_rate = round(
        (prev_total / prev_inspected * 100) if prev_inspected > 0 else 100.0, 1
    )

    prev_ts = sorted(ts for l in prev_conforming if (ts := _to_dt(l.timestamp)) is not None)
    prev_intervals: list[float] = []
    for i in range(1, len(prev_ts)):
        d = (prev_ts[i] - prev_ts[i - 1]).total_seconds()
        if 0 < d < 120:
            prev_intervals.append(d)
    avg_interval_prev = round(sum(prev_intervals) / len(prev_intervals), 2) if prev_intervals else 0.0

    # Deltas
    bags_delta_pct = round(((total_bags - prev_total) / prev_total * 100) if prev_total > 0 else 0.0, 1)
    interval_delta_pct = round(
        ((avg_interval - avg_interval_prev) / avg_interval_prev * 100) if avg_interval_prev > 0 else 0.0, 1
    )
    detection_rate_delta = round(detection_rate - prev_detection_rate, 1)

    # ── Sessions ─────────────────────────────────────────────────────────────
    sessions = db.query(models.Session).filter(models.Session.start_time >= start).all()
    active_sess = db.query(models.Session).filter(models.Session.status == "active").first()
    if active_sess and active_sess not in sessions:
        sessions.append(active_sess)

    total_session_seconds = 0.0
    for sess in sessions:
        s_start = _to_dt(sess.start_time)
        s_end   = _to_dt(sess.end_time) if sess.end_time else now
        if s_start is None:
            continue
        s_start = max(s_start, start)
        s_end   = min(s_end, now)
        dur = (s_end - s_start).total_seconds()
        if dur > 0:
            total_session_seconds += dur

    session_hours = round(total_session_seconds / 3600, 1)
    availability  = round(
        min(100.0, total_session_seconds / planned_seconds * 100) if planned_seconds > 0 else 0.0, 1
    )

    # OEE breakdown
    actual_rate = total_bags / (total_session_seconds / 3600) if total_session_seconds > 0 else 0.0
    performance = round(min(100.0, actual_rate / TARGET_RATE * 100) if TARGET_RATE > 0 else 0.0, 1)
    oee         = round((availability * performance * detection_rate) / 10000.0, 1)

    # Stop time formatted
    stop_secs = max(0.0, planned_seconds - total_session_seconds)
    stop_h = int(stop_secs // 3600)
    stop_m = int((stop_secs % 3600) // 60)
    stop_s = int(stop_secs % 60)
    stop_formatted = f"{stop_h:02d}:{stop_m:02d}:{stop_s:02d}"

    oee_data = [
        {"name": "Disponibilité", "value": availability, "color": "#22c55e"},
        {"name": "Performance",   "value": performance,  "color": "#f97316"},
        {"name": "Qualité",       "value": detection_rate, "color": "#3b82f6"},
    ]

    # ── Trend data (current vs previous, bucketed) ────────────────────────
    DAY_NAMES = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    if hours <= 24:
        bucket_size, bucket_count = 1, 24
    elif hours <= 168:
        bucket_size, bucket_count = 24, 7
    else:
        bucket_size, bucket_count = 168, max(1, hours // 168)

    trend_data = []
    for i in range(bucket_count - 1, -1, -1):
        b_start = now - _td(hours=(i + 1) * bucket_size)
        b_end   = now - _td(hours=i * bucket_size)
        pb_start = b_start - _td(hours=hours)
        pb_end   = b_end   - _td(hours=hours)

        if bucket_size >= 168:
            label = f"S{b_start.strftime('%W')}"
        elif bucket_size >= 24:
            label = DAY_NAMES[b_start.weekday()]
        else:
            label = b_start.strftime("%H:%M")

        cur_count = sum(
            1 for l in conforming
            if (ts := _to_dt(l.timestamp)) and b_start <= ts < b_end
        )
        prv_count = sum(
            1 for l in prev_conforming
            if (ts := _to_dt(l.timestamp)) and pb_start <= ts < pb_end
        )
        trend_data.append({"day": label, "current": cur_count, "previous": prv_count})

    # ── Key analyses ─────────────────────────────────────────────────────────
    peak_bucket = max(trend_data, key=lambda b: b["current"]) if trend_data else None

    consistency = 0.0
    if intervals and avg_interval > 0 and len(intervals) >= 2:
        variance    = sum((x - avg_interval) ** 2 for x in intervals) / len(intervals)
        cv          = _math.sqrt(variance) / avg_interval
        consistency = max(0.0, round((1.0 - cv) * 100.0, 1))

    return {
        "totalBags":          total_bags,
        "totalBagsPrev":      prev_total,
        "bagsDeltaPct":       bags_delta_pct,
        "avgInterval":        avg_interval,
        "avgIntervalPrev":    avg_interval_prev,
        "intervalDeltaPct":   interval_delta_pct,
        "detectionRate":      detection_rate,
        "detectionRatePrev":  prev_detection_rate,
        "detectionRateDelta": detection_rate_delta,
        "sessionHours":       session_hours,
        "availability":       availability,
        "oee":                oee,
        "performance":        performance,
        "stopTimeFormatted":  stop_formatted,
        "trendData":          trend_data,
        "oeeData":            oee_data,
        "consistency":        consistency,
        "peakBucket":         peak_bucket,
        "period":             period,
        "periodHours":        hours,
    }


# ─── Export System ───────────────────────────────────────────────────────────
_export_history: list[dict] = []

def _export_period_range(period: str, date_from: str = "", date_to: str = ""):
    """Return (start_dt, end_dt, period_label) for the requested period string.

    Aujourd'hui / Hier use calendar-day boundaries (UTC midnight) so that
    all bags counted during the day are included regardless of the current time.
    """
    from datetime import datetime as _dt, timedelta as _td
    now = _dt.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    if period == "yesterday":
        yesterday_start = today_start - _td(days=1)
        return yesterday_start, today_start, "Hier"
    if period == "last-7-days":
        return today_start - _td(days=7), now, "7 Derniers Jours"
    if period == "last-30-days":
        return today_start - _td(days=30), now, "30 Derniers Jours"
    if period == "custom" and date_from and date_to:
        try:
            return _dt.fromisoformat(date_from), _dt.fromisoformat(date_to), "Personnalisé"
        except Exception:
            pass
    # default: today (midnight → now)
    return today_start, now, "Aujourd'hui"


@app.get("/api/reports/export/preview")
async def export_preview(
    period: str = "today", source: str = "counts",
    date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db),
):
    """Return estimated row count and file size for the requested export."""
    start, end, _ = _export_period_range(period, date_from, date_to)
    AVG_BYTES = {"counts": 120, "sessions": 80, "anomalies": 100, "quality": 90}
    if source == "counts":
        count = db.query(models.DetectionLog).filter(
            models.DetectionLog.timestamp >= start,
            models.DetectionLog.timestamp <= end,
        ).count()
    elif source == "sessions":
        count = db.query(models.Session).filter(
            models.Session.start_time >= start,
            models.Session.start_time <= end,
        ).count()
    elif source == "anomalies":
        count = db.query(models.AlertHistory).filter(
            models.AlertHistory.timestamp >= start,
            models.AlertHistory.timestamp <= end,
        ).count()
    elif source == "quality":
        count = db.query(models.QualityReview).filter(
            models.QualityReview.created_at >= start,
            models.QualityReview.created_at <= end,
        ).count()
    else:
        count = 0
    size_kb = round((count * AVG_BYTES.get(source, 100)) / 1024, 1)
    return {"rows": count, "size_kb": size_kb}


@app.get("/api/reports/export/data")
async def export_data(
    period: str = "today", source: str = "counts", fmt: str = "csv",
    date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db),
):
    """Export data as CSV or JSON. Sources: counts | sessions | anomalies | quality."""
    from datetime import datetime as _dt
    from fastapi.responses import StreamingResponse
    import csv as _csv, io as _io, json as _json

    start, end, period_label = _export_period_range(period, date_from, date_to)
    ts_now = _dt.utcnow().strftime("%Y%m%d_%H%M%S")

    def _ts(v):
        return v.isoformat() if hasattr(v, "isoformat") else (str(v) if v else "")

    SOURCE_LABELS = {
        "counts": "Comptages Bruts", "sessions": "Sessions",
        "anomalies": "Anomalies-Alertes", "quality": "Qualité",
    }

    if source == "counts":
        rows_q = (db.query(models.DetectionLog)
                  .filter(models.DetectionLog.timestamp >= start,
                          models.DetectionLog.timestamp <= end)
                  .order_by(models.DetectionLog.timestamp.asc()).all())
        headers = ["ID", "Timestamp", "Session", "Statut", "Identifiant",
                   "Score Détection", "Score Logo", "Score Couleur", "Intervalle (s)", "Capture URL"]
        rows = [[l.id, _ts(l.timestamp), l.session_id, l.status, l.identifier or "",
                 round(l.detection_score or 0, 4), round(l.logo_score or 0, 4),
                 round(l.color_score or 0, 4), round(l.interval or 0, 3), l.capture_url or ""]
                for l in rows_q]
        fname_base = f"comptage_{period}_{ts_now}"
    elif source == "sessions":
        rows_q = (db.query(models.Session)
                  .filter(models.Session.start_time >= start,
                          models.Session.start_time <= end)
                  .order_by(models.Session.start_time.asc()).all())
        headers = ["ID", "Début", "Fin", "Total Sacs", "Sacs Rejetés", "Statut"]
        rows = [[s.id, _ts(s.start_time), _ts(s.end_time),
                 s.total_count or 0, s.rejected_count or 0, s.status]
                for s in rows_q]
        fname_base = f"sessions_{period}_{ts_now}"
    elif source == "anomalies":
        rows_q = (db.query(models.AlertHistory)
                  .filter(models.AlertHistory.timestamp >= start,
                          models.AlertHistory.timestamp <= end)
                  .order_by(models.AlertHistory.timestamp.asc()).all())
        headers = ["ID", "Timestamp", "Titre", "Message", "Type", "Lu"]
        rows = [[a.id, _ts(a.timestamp), a.title or "", a.message or "",
                 a.alert_type or "info", "oui" if a.is_read else "non"]
                for a in rows_q]
        fname_base = f"alertes_{period}_{ts_now}"
    elif source == "quality":
        rows_q = (db.query(models.QualityReview)
                  .filter(models.QualityReview.created_at >= start,
                          models.QualityReview.created_at <= end)
                  .order_by(models.QualityReview.created_at.asc()).all())
        headers = ["ID", "Log ID", "Action", "Statut Cible", "Notes", "Réviseur", "Date"]
        rows = [[r.id, r.log_id, r.action, r.target_status or "",
                 r.notes or "", r.reviewer or "", _ts(r.created_at)]
                for r in rows_q]
        fname_base = f"qualite_{period}_{ts_now}"
    else:
        raise HTTPException(status_code=400, detail="Source invalide")

    # ── Record in history ─────────────────────────────────────────────────
    _export_history.insert(0, {
        "name": f"{fname_base}.{fmt}",
        "source": SOURCE_LABELS.get(source, source),
        "period_label": period_label,
        "rows": len(rows),
        "size_kb": round((len(rows) * 100) / 1024, 1),
        "timestamp": _dt.utcnow().isoformat(),
        "format": fmt,
    })
    if len(_export_history) > 20:
        _export_history.pop()

    # ── Generate output ───────────────────────────────────────────────────
    if fmt == "json":
        data = [dict(zip(headers, r)) for r in rows]
        content = _json.dumps(
            {"period": period_label, "source": source, "data": data},
            ensure_ascii=False, indent=2,
        )
        return StreamingResponse(
            iter([content]),
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={fname_base}.json"},
        )

    elif fmt == "xlsx":
        import openpyxl as _xl
        from openpyxl.styles import Font as _XF, PatternFill as _XP, Alignment as _XA
        wb = _xl.Workbook()
        ws = wb.active
        ws.title = SOURCE_LABELS.get(source, source)[:31]
        # Title row (merged across all columns)
        last_col_letter = chr(64 + min(len(headers), 26))
        ws.merge_cells(f"A1:{last_col_letter}1")
        tc = ws["A1"]
        tc.value = f"Export — {SOURCE_LABELS.get(source, source)} — {period_label}"
        tc.font = _XF(bold=True, size=12, color="FFFFFF")
        tc.fill = _XP(start_color="E85D04", end_color="E85D04", fill_type="solid")
        tc.alignment = _XA(horizontal="center")
        # Info row
        ws.append(["Généré:", _dt.utcnow().strftime("%d/%m/%Y %H:%M UTC"),
                   "Période:", period_label, "Lignes:", len(rows)])
        for cell in ws[2]:
            cell.font = _XF(color="999999", italic=True, size=8)
        # Header row (row 3)
        ws.append(headers)
        for cell in ws[3]:
            cell.font = _XF(bold=True, color="FFFFFF")
            cell.fill = _XP(start_color="222222", end_color="222222", fill_type="solid")
            cell.alignment = _XA(horizontal="center")
        # Data rows
        for i, row in enumerate(rows):
            ws.append(row)
            row_fill = "181818" if i % 2 == 0 else "242424"
            for cell in ws[i + 4]:
                cell.fill = _XP(start_color=row_fill, end_color=row_fill, fill_type="solid")
        # Auto-width — skip MergedCell objects which have no column_letter
        for col in ws.columns:
            real_cells = [c for c in col if hasattr(c, "column_letter")]
            if not real_cells:
                continue
            max_len = max((len(str(c.value or "")) for c in real_cells), default=10)
            ws.column_dimensions[real_cells[0].column_letter].width = min(max_len + 2, 50)
        xlsx_buf = _io.BytesIO()
        wb.save(xlsx_buf)
        xlsx_buf.seek(0)
        return StreamingResponse(
            iter([xlsx_buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname_base}.xlsx"},
        )

    elif fmt == "pdf":
        from fpdf import FPDF as _FPDF

        def _ps(s: str) -> str:
            """Sanitize string for fpdf2 latin-1 built-in fonts."""
            return (s.replace("\u2014", "-").replace("\u2013", "-")
                     .replace("\u2018", "'").replace("\u2019", "'")
                     .replace("\u201c", '"').replace("\u201d", '"')
                     .encode("latin-1", errors="replace").decode("latin-1"))

        pdf = _FPDF(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_margins(10, 10, 10)
        pdf.add_page()
        # Title
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_text_color(232, 93, 4)
        pdf.cell(0, 10, _ps(f"Rapport Export - {SOURCE_LABELS.get(source, source)}"), ln=True)
        # Subtitle
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(150, 150, 150)
        info = _ps(f"Periode : {period_label}   |   "
                   f"Genere le {_dt.utcnow().strftime('%d/%m/%Y a %H:%M')} UTC   |   "
                   f"{len(rows)} enregistrement(s)")
        pdf.cell(0, 6, info, ln=True)
        pdf.ln(3)
        # Table layout (A4 landscape usable ≈ 277 mm)
        n_cols = len(headers)
        col_w = 277 / n_cols
        # Header row
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(45, 45, 45)
        pdf.set_text_color(232, 93, 4)
        pdf.set_draw_color(80, 80, 80)
        for h in headers:
            pdf.cell(col_w, 6, _ps(str(h)[:22]), border=1, fill=True, align="C")
        pdf.ln()
        # Data rows (max 2 000 to keep PDF manageable)
        pdf.set_font("Helvetica", "", 6)
        for i, row in enumerate(rows[:2000]):
            r, g, b = (255, 255, 255) if i % 2 == 0 else (242, 242, 242)
            pdf.set_fill_color(r, g, b)
            pdf.set_text_color(40, 40, 40)
            for val in row:
                pdf.cell(col_w, 5, _ps(str(val)[:22]), border=1, fill=True)
            pdf.ln()
        if len(rows) > 2000:
            pdf.set_text_color(200, 80, 80)
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(0, 8,
                     _ps(f"Avertissement : PDF tronque a 2 000 lignes (total : {len(rows)}). "
                         "Utilisez CSV ou XLSX pour l'export complet."), ln=True)
        pdf_bytes = bytes(pdf.output())
        return StreamingResponse(
            iter([pdf_bytes]),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={fname_base}.pdf"},
        )

    else:  # csv (default)
        buf = _io.StringIO()
        writer = _csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={fname_base}.csv"},
        )


@app.get("/api/reports/export/history")
async def get_export_history():
    """Return the list of recent in-memory export records (max 20)."""
    return _export_history


# Keep the old CSV route as a redirect-compatible alias
@app.get("/api/reports/export/csv")
async def export_production_csv_legacy(period: str = "week", db: Session = Depends(get_db)):
    """Legacy CSV export alias — maps to new /api/reports/export/data."""
    PERIOD_MAP = {"day": "today", "week": "last-7-days", "month": "last-30-days"}
    return await export_data(period=PERIOD_MAP.get(period, "last-7-days"), source="counts", fmt="csv", db=db)


# ─── Scheduled Exports ───────────────────────────────────────────────────────
import os as _os
from pathlib import Path as _Path

_SCHED_EXPORT_DIR = _Path("backend/static/exports")
_SCHED_KEYS = [
    "export_sched_enabled", "export_sched_frequency", "export_sched_time",
    "export_sched_day_of_week", "export_sched_day_of_month",
    "export_sched_source", "export_sched_format", "export_sched_period", "export_sched_email",
]
_SCHED_DEFAULTS = {
    "export_sched_enabled":      "false",
    "export_sched_frequency":    "daily",
    "export_sched_time":         "06:00",
    "export_sched_day_of_week":  "1",
    "export_sched_day_of_month": "1",
    "export_sched_source":       "counts",
    "export_sched_format":       "csv",
    "export_sched_period":       "yesterday",
    "export_sched_email":        "",
}
_schedule_task: asyncio.Task | None = None
_scheduled_exports: list[dict] = []
_last_sched_minute: str = ""


def _get_sched_config(db) -> dict:
    rows = db.query(models.SystemSetting).filter(
        models.SystemSetting.key.in_(_SCHED_KEYS)
    ).all()
    cfg = dict(_SCHED_DEFAULTS)
    cfg.update({r.key: r.value for r in rows})
    return {
        "enabled":        cfg["export_sched_enabled"] == "true",
        "frequency":      cfg["export_sched_frequency"],
        "time":           cfg["export_sched_time"],
        "day_of_week":    int(cfg["export_sched_day_of_week"]),
        "day_of_month":   int(cfg["export_sched_day_of_month"]),
        "source":         cfg["export_sched_source"],
        "format":         cfg["export_sched_format"],
        "period":         cfg["export_sched_period"],
        "email":          cfg["export_sched_email"],
    }


def _save_sched_config(db, payload: dict):
    mapping = {
        "enabled":        ("export_sched_enabled",      lambda v: "true" if v else "false"),
        "frequency":      ("export_sched_frequency",    str),
        "time":           ("export_sched_time",         str),
        "day_of_week":    ("export_sched_day_of_week",  str),
        "day_of_month":   ("export_sched_day_of_month", str),
        "source":         ("export_sched_source",       str),
        "format":         ("export_sched_format",       str),
        "period":         ("export_sched_period",       str),
        "email":          ("export_sched_email",        str),
    }
    for field, (key, conv) in mapping.items():
        if field in payload:
            row = db.query(models.SystemSetting).filter_by(key=key).first()
            val = conv(payload[field])
            if row:
                row.value = val
            else:
                db.add(models.SystemSetting(key=key, value=val))
    db.commit()


def _do_scheduled_export(cfg: dict):
    """Build export bytes and save to SCHED_EXPORT_DIR. Returns filename."""
    from datetime import datetime as _dt
    import csv as _csv, io as _io, json as _json

    _SCHED_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    source = cfg["source"]
    fmt    = cfg["format"]
    period = cfg["period"]
    start, end, period_label = _export_period_range(period)
    ts_now = _dt.utcnow().strftime("%Y%m%d_%H%M%S")

    def _ts(v):
        return v.isoformat() if hasattr(v, "isoformat") else (str(v) if v else "")

    SOURCE_LABELS = {
        "counts": "Comptages Bruts", "sessions": "Sessions",
        "anomalies": "Anomalies-Alertes", "quality": "Qualité",
    }

    db = SessionLocal()
    try:
        if source == "counts":
            rows_q = (db.query(models.DetectionLog)
                      .filter(models.DetectionLog.timestamp >= start,
                              models.DetectionLog.timestamp <= end)
                      .order_by(models.DetectionLog.timestamp.asc()).all())
            headers = ["ID", "Timestamp", "Session", "Statut", "Identifiant",
                       "Score Détection", "Score Logo", "Score Couleur", "Intervalle (s)", "Capture URL"]
            rows = [[l.id, _ts(l.timestamp), l.session_id, l.status, l.identifier or "",
                     round(l.detection_score or 0, 4), round(l.logo_score or 0, 4),
                     round(l.color_score or 0, 4), round(l.interval or 0, 3), l.capture_url or ""]
                    for l in rows_q]
            fname_base = f"auto_comptage_{period}_{ts_now}"
        elif source == "sessions":
            rows_q = (db.query(models.Session)
                      .filter(models.Session.start_time >= start,
                              models.Session.start_time <= end)
                      .order_by(models.Session.start_time.asc()).all())
            headers = ["ID", "Début", "Fin", "Total Sacs", "Sacs Rejetés", "Statut"]
            rows = [[s.id, _ts(s.start_time), _ts(s.end_time),
                     s.total_count or 0, s.rejected_count or 0, s.status]
                    for s in rows_q]
            fname_base = f"auto_sessions_{period}_{ts_now}"
        elif source == "anomalies":
            rows_q = (db.query(models.AlertHistory)
                      .filter(models.AlertHistory.timestamp >= start,
                              models.AlertHistory.timestamp <= end)
                      .order_by(models.AlertHistory.timestamp.asc()).all())
            headers = ["ID", "Timestamp", "Titre", "Message", "Type", "Lu"]
            rows = [[a.id, _ts(a.timestamp), a.title or "", a.message or "",
                     a.alert_type or "info", "oui" if a.is_read else "non"]
                    for a in rows_q]
            fname_base = f"auto_alertes_{period}_{ts_now}"
        else:  # quality
            rows_q = (db.query(models.QualityReview)
                      .filter(models.QualityReview.created_at >= start,
                              models.QualityReview.created_at <= end)
                      .order_by(models.QualityReview.created_at.asc()).all())
            headers = ["ID", "Log ID", "Action", "Statut Cible", "Notes", "Réviseur", "Date"]
            rows = [[r.id, r.log_id, r.action, r.target_status or "",
                     r.notes or "", r.reviewer or "", _ts(r.created_at)]
                    for r in rows_q]
            fname_base = f"auto_qualite_{period}_{ts_now}"
    finally:
        db.close()

    fname = f"{fname_base}.{fmt}"
    fpath = _SCHED_EXPORT_DIR / fname

    if fmt == "json":
        data = [dict(zip(headers, r)) for r in rows]
        content = _json.dumps({"period": period_label, "source": source, "data": data},
                              ensure_ascii=False, indent=2)
        fpath.write_text(content, encoding="utf-8")
    elif fmt == "xlsx":
        import openpyxl as _xl
        from openpyxl.styles import Font as _XF, PatternFill as _XP, Alignment as _XA
        wb = _xl.Workbook()
        ws = wb.active
        ws.title = SOURCE_LABELS.get(source, source)[:31]
        last_col = chr(64 + min(len(headers), 26))
        ws.merge_cells(f"A1:{last_col}1")
        tc = ws["A1"]
        tc.value = f"Export — {SOURCE_LABELS.get(source, source)} — {period_label}"
        tc.font = _XF(bold=True, size=12, color="FFFFFF")
        tc.fill = _XP(start_color="E85D04", end_color="E85D04", fill_type="solid")
        tc.alignment = _XA(horizontal="center")
        ws.append(["Généré:", _dt.utcnow().strftime("%d/%m/%Y %H:%M UTC"), "Période:", period_label])
        ws.append(headers)
        for cell in ws[3]:
            cell.font = _XF(bold=True, color="FFFFFF")
            cell.fill = _XP(start_color="222222", end_color="222222", fill_type="solid")
        for i, row in enumerate(rows):
            ws.append(row)
        wb.save(str(fpath))
    elif fmt == "pdf":
        from fpdf import FPDF as _FPDF

        def _ps(s: str) -> str:
            return (s.replace("\u2014", "-").replace("\u2013", "-")
                     .replace("\u2018", "'").replace("\u2019", "'")
                     .replace("\u201c", '"').replace("\u201d", '"')
                     .encode("latin-1", errors="replace").decode("latin-1"))

        pdf = _FPDF(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_margins(10, 10, 10)
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_text_color(232, 93, 4)
        pdf.cell(0, 10, _ps(f"Rapport Export - {SOURCE_LABELS.get(source, source)}"), ln=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 6, _ps(f"Periode : {period_label}   |   {_dt.utcnow().strftime('%d/%m/%Y %H:%M')} UTC   |   {len(rows)} enreg."), ln=True)
        pdf.ln(3)
        col_w = 277 / len(headers)
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(45, 45, 45)
        pdf.set_text_color(232, 93, 4)
        for h in headers:
            pdf.cell(col_w, 6, _ps(str(h)[:22]), border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 6)
        for i, row in enumerate(rows[:2000]):
            r, g, b = (255, 255, 255) if i % 2 == 0 else (242, 242, 242)
            pdf.set_fill_color(r, g, b)
            pdf.set_text_color(40, 40, 40)
            for val in row:
                pdf.cell(col_w, 5, _ps(str(val)[:22]), border=1, fill=True)
            pdf.ln()
        fpath.write_bytes(bytes(pdf.output()))
    else:  # csv
        buf = _io.StringIO()
        writer = _csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        fpath.write_text(buf.getvalue(), encoding="utf-8")

    size_kb = round(fpath.stat().st_size / 1024, 1) if fpath.exists() else 0
    _scheduled_exports.insert(0, {
        "name": fname,
        "source": SOURCE_LABELS.get(source, source),
        "period_label": period_label,
        "rows": len(rows),
        "size_kb": size_kb,
        "triggered_at": _dt.utcnow().isoformat(),
        "format": fmt,
        "download_url": f"/static/exports/{fname}",
    })
    if len(_scheduled_exports) > 10:
        _scheduled_exports.pop()

    # ── Envoi par email si configuré ─────────────────────────────────────────
    email_to = cfg.get("email", "").strip()
    if email_to:
        _db_mail = SessionLocal()
        try:
            smtp_cfg = _get_smtp_config(_db_mail)
            if smtp_cfg["host"] and smtp_cfg["user"]:
                _send_email(
                    smtp_cfg,
                    to=email_to,
                    subject=f"[CimentMonitor] Rapport planifié — {SOURCE_LABELS.get(source, source)} ({period_label})",
                    body=(
                        f"Bonjour,\n\n"
                        f"Veuillez trouver en pièce jointe le rapport planifié :\n"
                        f"  • Source  : {SOURCE_LABELS.get(source, source)}\n"
                        f"  • Période : {period_label}\n"
                        f"  • Lignes  : {len(rows)}\n"
                        f"  • Taille  : {size_kb} KB\n\n"
                        f"— CimentMonitor Pro"
                    ),
                    attachment_path=str(fpath),
                )
        except Exception as _mail_err:
            logger.warning("Échec envoi email rapport planifié : %s", _mail_err)
        finally:
            _db_mail.close()

    return fname


async def _scheduler_loop():
    global _last_sched_minute
    while True:
        try:
            await asyncio.sleep(30)
            db = SessionLocal()
            try:
                cfg = _get_sched_config(db)
            finally:
                db.close()
            if not cfg["enabled"]:
                continue
            from datetime import datetime as _dt
            now = _dt.now()   # heure locale du serveur (pas UTC)
            try:
                th, tm = map(int, cfg["time"].split(":"))
            except Exception:
                continue
            if now.hour != th or now.minute != tm:
                continue
            minute_key = f"{now.date().isoformat()}T{th:02d}:{tm:02d}"
            if minute_key == _last_sched_minute:
                continue
            freq = cfg["frequency"]
            if freq == "weekly" and now.isoweekday() != cfg["day_of_week"]:
                continue
            if freq == "monthly" and now.day != cfg["day_of_month"]:
                continue
            _last_sched_minute = minute_key
            _do_scheduled_export(cfg)
        except asyncio.CancelledError:
            break
        except Exception:
            pass


@app.get("/api/reports/export/schedule")
async def get_schedule(db: Session = Depends(get_db)):
    return _get_sched_config(db)


@app.put("/api/reports/export/schedule")
async def update_schedule(payload: dict, db: Session = Depends(get_db)):
    _save_sched_config(db, payload)
    return _get_sched_config(db)


@app.post("/api/reports/export/schedule/run")
async def run_schedule_now(db: Session = Depends(get_db)):
    """Manually trigger a scheduled export immediately."""
    cfg = _get_sched_config(db)
    try:
        fname = _do_scheduled_export(cfg)
        return {"status": "ok", "file": fname}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/reports/export/scheduled")
async def get_scheduled_exports():
    return _scheduled_exports


# ─── Timeline ────────────────────────────────────────────────────────────────
@app.get("/api/timeline/hourly")
async def get_timeline_hourly(hours: int = 24, db: Session = Depends(get_db)):
    """Distribution horaire de production sur les N dernières heures (max 168h)."""
    from datetime import datetime as _dt, timedelta as _td

    hours = max(1, min(hours, 168))
    now = _dt.utcnow()
    start = now - _td(hours=hours)

    logs = (
        db.query(models.DetectionLog)
        .filter(models.DetectionLog.timestamp >= start)
        .order_by(models.DetectionLog.timestamp.asc())
        .all()
    )

    # Normalisation timestamp
    def _to_dt(ts):
        if isinstance(ts, str):
            try:
                return _dt.fromisoformat(ts)
            except Exception:
                return None
        return ts

    # Groupement par bucket horaire
    data = []
    for i in range(hours - 1, -1, -1):
        b_start = now - _td(hours=i + 1)
        b_end   = now - _td(hours=i)
        label   = b_start.strftime("%H:%M")

        b_conformes = []
        b_rejected  = 0
        for log in logs:
            ts = _to_dt(log.timestamp)
            if ts is None or not (b_start <= ts < b_end):
                continue
            if log.status == "conforme":
                b_conformes.append(ts)
            elif log.status == "rejete":
                b_rejected += 1

        # Intervalle moyen dans le bucket
        iv_list = []
        for j in range(1, len(b_conformes)):
            delta = (b_conformes[j] - b_conformes[j - 1]).total_seconds()
            if 0 < delta < 120:
                iv_list.append(delta)

        data.append({
            "time":     label,
            "count":    len(b_conformes),
            "interval": round(sum(iv_list) / len(iv_list), 2) if iv_list else 0.0,
            "rejected": b_rejected,
        })

    # Analyse des pics (seulement les buckets non vides)
    non_empty = [b for b in data if b["count"] > 0]
    peak_max  = max(non_empty, key=lambda b: b["count"]) if non_empty else None
    peak_min  = min(non_empty, key=lambda b: b["count"]) if non_empty else None

    return {
        "data":        data,
        "peakMax":     peak_max,
        "peakMin":     peak_min,
        "totalBags":   sum(b["count"]    for b in data),
        "totalRejected": sum(b["rejected"] for b in data),
        "periodHours": hours,
    }


# ─── Dashboard ────────────────────────────────────────────────────────────────
@app.get("/api/dashboard/summary")
async def get_dashboard_summary(db: Session = Depends(get_db)):
    import math as _math
    from datetime import datetime as _dt, timedelta as _td

    now = _dt.utcnow()

    active_session = db.query(models.Session).filter(models.Session.status == "active").first()
    total_bags = db.query(models.DetectionLog).filter(models.DetectionLog.status == "conforme").count()
    rejected_bags = db.query(models.DetectionLog).filter(models.DetectionLog.status == "rejete").count()

    # Logs conformes de la session active, triés chronologiquement
    if active_session:
        session_logs = (
            db.query(models.DetectionLog)
            .filter(
                models.DetectionLog.session_id == active_session.id,
                models.DetectionLog.status == "conforme",
            )
            .order_by(models.DetectionLog.timestamp.asc())
            .all()
        )
    else:
        session_logs = []

    # Normalisation des timestamps (string ISO ou datetime)
    timestamps = []
    for log in session_logs:
        ts = log.timestamp
        if isinstance(ts, str):
            try:
                ts = _dt.fromisoformat(ts)
            except Exception:
                continue
        timestamps.append(ts)

    # Intervalles entre sacs consécutifs (on ignore les pauses > 120s)
    ts_intervals: list[tuple] = []  # (timestamp_du_sac_précédent, durée)
    for i in range(1, len(timestamps)):
        delta = (timestamps[i] - timestamps[i - 1]).total_seconds()
        if 0 < delta < 120:
            ts_intervals.append((timestamps[i - 1], delta))

    all_iv = [iv for _, iv in ts_intervals]
    avg_interval = sum(all_iv) / len(all_iv) if all_iv else 0.0

    # Taux de production : sacs dans les 5 dernières minutes
    cutoff_5min = now - _td(minutes=5)
    recent_count = sum(1 for ts in timestamps if ts >= cutoff_5min)
    production_rate = round(recent_count / 5.0, 2)

    # Consistance : 100 × (1 − CV), CV = σ/μ
    if len(all_iv) >= 2 and avg_interval > 0:
        variance = sum((x - avg_interval) ** 2 for x in all_iv) / len(all_iv)
        stddev = _math.sqrt(variance)
        cv = stddev / avg_interval
        consistency = max(0.0, round((1.0 - cv) * 100.0, 1))
        stddev_val = round(stddev, 2)
    else:
        stddev_val = 0.0
        consistency = 100.0

    # Première / deuxième moitié de session
    if len(all_iv) >= 4:
        mid = len(all_iv) // 2
        fh = all_iv[:mid]
        sh = all_iv[mid:]
        first_half_interval = round(sum(fh) / len(fh), 2)
        second_half_interval = round(sum(sh) / len(sh), 2)
        slowdown_pct = (
            round(((second_half_interval - first_half_interval) / first_half_interval) * 100.0, 1)
            if first_half_interval > 0 else 0.0
        )
    else:
        first_half_interval = round(avg_interval, 2)
        second_half_interval = round(avg_interval, 2)
        slowdown_pct = 0.0

    # Graphique intervalles : buckets par minute sur les 14 dernières minutes
    interval_data = []
    for i in range(13, -1, -1):
        b_start = now - _td(minutes=i + 1)
        b_end = now - _td(minutes=i)
        label = f"{b_start.hour}:{str(b_start.minute).zfill(2)}"
        b_iv = [iv for ts, iv in ts_intervals if b_start <= ts < b_end]
        if b_iv:
            interval_data.append({
                "time": label,
                "avgInterval": round(sum(b_iv) / len(b_iv), 2),
                "minInterval": round(min(b_iv), 2),
                "maxInterval": round(max(b_iv), 2),
            })
        else:
            interval_data.append({"time": label, "avgInterval": 0, "minInterval": 0, "maxInterval": 0})

    # Heatmap : 6 buckets de 5 secondes (dernières 30s)
    heatmap_data = []
    for i in range(5, -1, -1):
        b_start = now - _td(seconds=(i + 1) * 5)
        b_end = now - _td(seconds=i * 5)
        count = sum(1 for ts in timestamps if b_start <= ts < b_end)
        level = "none" if count == 0 else "low" if count < 2 else "medium" if count < 4 else "high"
        heatmap_data.append({"time": f"{(5 - i) * 5}s", "activity": {"level": level, "count": count}})

    # Production gaps : intervalles > 2× la moyenne
    production_gaps = []
    if avg_interval > 0 and len(timestamps) >= 2:
        for i in range(1, len(timestamps)):
            delta = (timestamps[i] - timestamps[i - 1]).total_seconds()
            if delta > 2 * avg_interval:
                deviation_pct = int(((delta - avg_interval) / avg_interval) * 100)
                production_gaps.append({
                    "id": str(i),
                    "bagRange": f"#{i} → #{i + 1}",
                    "duration": f"{delta:.2f}s",
                    "time": timestamps[i - 1].strftime("%H:%M"),
                    "deviation": deviation_pct,
                })

    return {
        "totalBags": total_bags,
        "rejectedBags": rejected_bags,
        "activeSessionId": active_session.id if active_session else None,
        "productionRate": production_rate,
        "avgInterval": round(avg_interval, 2),
        "consistency": consistency,
        "stddev": stddev_val,
        "firstHalfInterval": first_half_interval,
        "secondHalfInterval": second_half_interval,
        "slowdownPercent": slowdown_pct,
        "intervalData": interval_data,
        "heatmapData": heatmap_data,
        "productionGaps": production_gaps,
    }


# ─── Logs ─────────────────────────────────────────────────────────────────────
@app.get("/api/logs/", response_model=schemas.DetectionLogListResponse)
async def get_logs(
    page: int = 1,
    page_size: int = 20,
    status: str | None = None,
    session_id: str | None = None,
    search: str | None = None,
    db: Session = Depends(get_db),
):
    page = max(page, 1)
    page_size = min(max(page_size, 1), 100)

    query = db.query(models.DetectionLog)
    if status:
        mapped_status = "conforme" if status.lower() in ["verifie", "vérifié", "conforme"] else "rejete" if status.lower() in ["rejete", "rejeté"] else status
        query = query.filter(models.DetectionLog.status == mapped_status)
    if session_id:
        query = query.filter(models.DetectionLog.session_id == session_id)
    if search:
        query = query.filter(models.DetectionLog.identifier.ilike(f"%{search}%"))

    total = query.count()
    logs = query.order_by(models.DetectionLog.timestamp.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": logs,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ─── Users ────────────────────────────────────────────────────────────────────
import datetime as _dt_users

def _user_to_dict(u: models.User) -> dict:
    return {
        "id": u.id,
        "username": u.username,
        "full_name": u.full_name,
        "role": u.role,
        "is_active": u.is_active,
        "last_login": u.last_login.isoformat() if u.last_login else None,
        "login_count": u.login_count or 0,
    }

# ─── Permissions catalogue ────────────────────────────────────────────────────

PERMISSIONS_CATALOG = [
    # Dashboard
    {"id": "dashboard_view",    "group": "Tableau de Bord",  "label": "Voir le tableau de bord"},
    # Monitoring
    {"id": "livestream_view",   "group": "Monitoring",       "label": "Flux vidéo en direct"},
    # Production
    {"id": "sessions_manage",   "group": "Production",       "label": "Gérer les sessions (démarrer/arrêter)"},
    {"id": "logs_view",         "group": "Production",       "label": "Voir les logs de production"},
    {"id": "timeline_view",     "group": "Production",       "label": "Voir la chronologie"},
    # Configuration
    {"id": "config_camera",     "group": "Configuration",    "label": "Paramètres caméra"},
    {"id": "config_model",      "group": "Configuration",    "label": "Modèle IA"},
    {"id": "config_templates",  "group": "Configuration",    "label": "Templates & Couleurs"},
    {"id": "config_line",       "group": "Configuration",    "label": "Ligne virtuelle"},
    # Qualité
    {"id": "quality_view",      "group": "Qualité",          "label": "Tableau de bord qualité"},
    {"id": "anomalies_view",    "group": "Qualité",          "label": "Détection d'anomalies"},
    # Alertes
    {"id": "alerts_view",       "group": "Alertes",          "label": "Voir les alertes"},
    {"id": "alerts_manage",     "group": "Alertes",          "label": "Gérer les alertes (règles/seuils)"},
    # Rapports
    {"id": "reports_view",      "group": "Rapports",         "label": "Voir les rapports de production"},
    {"id": "reports_export",    "group": "Rapports",         "label": "Exporter les données (CSV/PDF/XLSX)"},
    # Analytique
    {"id": "analytics_view",    "group": "Analytique",       "label": "Performance & OEE"},
    # Administration
    {"id": "users_manage",      "group": "Administration",   "label": "Gérer les utilisateurs & rôles"},
    {"id": "system_settings",   "group": "Administration",   "label": "Paramètres système"},
    {"id": "devices_manage",    "group": "Administration",   "label": "Gérer les appareils & caméras"},
    # Maintenance
    {"id": "maintenance_view",  "group": "Maintenance",      "label": "Santé système"},
    {"id": "database_manage",   "group": "Maintenance",      "label": "Gestion de la base de données"},
]

_ALL_PERMS = [p["id"] for p in PERMISSIONS_CATALOG]

_DEFAULT_ROLES_SEED = [
    {
        "name": "admin", "label": "Administrateur",
        "description": "Accès complet au système. Peut configurer le modèle IA, gérer les utilisateurs et modifier tous les paramètres.",
        "permissions": _ALL_PERMS, "is_builtin": True,
    },
    {
        "name": "operator", "label": "Opérateur",
        "description": "Peut démarrer/arrêter des sessions de production, surveiller la qualité et créer des alertes manuelles.",
        "permissions": [
            "dashboard_view", "livestream_view",
            "sessions_manage", "logs_view", "timeline_view",
            "quality_view", "anomalies_view",
            "alerts_view", "alerts_manage",
            "reports_view", "reports_export",
            "analytics_view", "maintenance_view",
        ],
        "is_builtin": True,
    },
    {
        "name": "viewer", "label": "Lecteur",
        "description": "Accès en lecture seule aux tableaux de bord et rapports. Aucune action de modification possible.",
        "permissions": [
            "dashboard_view", "livestream_view",
            "logs_view", "timeline_view",
            "quality_view", "reports_view", "analytics_view",
        ],
        "is_builtin": True,
    },
]


# ─── Users & Roles ─────────────────────────────────────────────────────────────

def _require_admin(current_user: models.User = Depends(auth.get_current_user)) -> models.User:
    """Dependency: rejects non-admin users with 403."""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    return current_user


@app.get("/api/users/")
async def get_users(
    db: Session = Depends(get_db),
    _: models.User = Depends(_require_admin),
):
    return [_user_to_dict(u) for u in db.query(models.User).order_by(models.User.id).all()]


@app.post("/api/users/", status_code=201)
async def create_user(
    payload: schemas.UserCreate,
    db: Session = Depends(get_db),
    _: models.User = Depends(_require_admin),
):
    if db.query(models.User).filter(models.User.username == payload.username).first():
        raise HTTPException(status_code=409, detail="Nom d'utilisateur déjà pris")
    user = models.User(
        username=payload.username,
        full_name=payload.full_name,
        role=payload.role,
        hashed_password=auth.get_password_hash(payload.password),
        is_active=True,
        login_count=0,
    )
    db.add(user)
    db.flush()
    db.add(models.UserActivity(user_id=user.id, username=user.username, action="created"))
    db.commit()
    db.refresh(user)
    return _user_to_dict(user)


@app.put("/api/users/{user_id}")
async def update_user(
    user_id: int,
    payload: schemas.UserUpdate,
    db: Session = Depends(get_db),
    _: models.User = Depends(_require_admin),
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")
    for field, val in payload.model_dump(exclude_unset=True).items():
        setattr(user, field, val)
    db.add(models.UserActivity(user_id=user.id, username=user.username, action="updated"))
    db.commit()
    db.refresh(user)
    return _user_to_dict(user)


@app.patch("/api/users/{user_id}/password")
async def change_password(
    user_id: int,
    payload: schemas.UserPasswordChange,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    # Un admin peut changer le mdp de n'importe qui ; un user ne peut changer que le sien
    if current_user.role != "admin" and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Vous ne pouvez modifier que votre propre mot de passe")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")
    if len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="Le mot de passe doit contenir au moins 6 caractères")
    user.hashed_password = auth.get_password_hash(payload.new_password)
    db.add(models.UserActivity(user_id=user.id, username=user.username, action="password_changed"))
    db.commit()
    return {"ok": True}


@app.delete("/api/users/{user_id}")
async def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(_require_admin),
):
    if current_user.id == user_id:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas supprimer votre propre compte")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")
    uname = user.username
    db.delete(user)
    db.add(models.UserActivity(user_id=None, username=uname, action="deleted"))
    db.commit()
    return {"ok": True}


@app.get("/api/users/activity")
async def get_user_activity(
    limit: int = 50,
    page: int = 1,
    db: Session = Depends(get_db),
    _: models.User = Depends(_require_admin),
):
    offset = (page - 1) * limit
    total = db.query(models.UserActivity).count()
    rows = (
        db.query(models.UserActivity)
        .order_by(models.UserActivity.timestamp.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    return {
        "total": total,
        "page": page,
        "page_size": limit,
        "items": [
            {
                "id": r.id,
                "username": r.username,
                "timestamp": r.timestamp.isoformat(),
                "action": r.action,
                "ip_address": r.ip_address,
                "user_agent": r.user_agent,
            }
            for r in rows
        ],
    }


# ─── Roles ────────────────────────────────────────────────────────────────────

def _role_to_dict(r: models.Role, db) -> dict:
    import json as _j
    perms = _j.loads(r.permissions or "[]")
    user_count = db.query(models.User).filter(models.User.role == r.name).count()
    return {
        "id": r.id, "name": r.name, "label": r.label,
        "description": r.description, "permissions": perms,
        "is_builtin": r.is_builtin, "user_count": user_count,
    }


@app.get("/api/roles/permissions")
async def list_permission_catalog(_: models.User = Depends(auth.get_current_user)):
    """Return the full catalogue of available permission slugs, grouped by module."""
    return PERMISSIONS_CATALOG


@app.get("/api/roles/")
async def list_roles(
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    roles = db.query(models.Role).order_by(models.Role.id).all()
    return [_role_to_dict(r, db) for r in roles]


@app.post("/api/roles/", status_code=201)
async def create_role(
    payload: schemas.RoleCreate,
    db: Session = Depends(get_db),
    _: models.User = Depends(_require_admin),
):
    import json as _j
    import re as _re
    if not _re.match(r'^[a-z0-9_]+$', payload.name):
        raise HTTPException(status_code=400, detail="Le nom de rôle doit être en minuscules sans espaces (ex: chef_equipe)")
    if db.query(models.Role).filter(models.Role.name == payload.name).first():
        raise HTTPException(status_code=409, detail="Un rôle avec ce nom existe déjà")
    db_role = models.Role(
        name=payload.name, label=payload.label,
        description=payload.description,
        permissions=_j.dumps(payload.permissions),
        is_builtin=False,
    )
    db.add(db_role)
    db.commit()
    db.refresh(db_role)
    return _role_to_dict(db_role, db)


@app.put("/api/roles/{role_id}")
async def update_role(
    role_id: int,
    payload: schemas.RoleUpdate,
    db: Session = Depends(get_db),
    _: models.User = Depends(_require_admin),
):
    import json as _j
    db_role = db.query(models.Role).filter(models.Role.id == role_id).first()
    if not db_role:
        raise HTTPException(status_code=404, detail="Rôle introuvable")
    if payload.label is not None:
        if db_role.is_builtin:
            raise HTTPException(status_code=400, detail="Le libellé des rôles intégrés ne peut pas être modifié")
        db_role.label = payload.label
    if payload.description is not None:
        db_role.description = payload.description
    if payload.permissions is not None:
        db_role.permissions = _j.dumps(payload.permissions)
    db.commit()
    db.refresh(db_role)
    return _role_to_dict(db_role, db)


@app.delete("/api/roles/{role_id}")
async def delete_role(
    role_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(_require_admin),
):
    db_role = db.query(models.Role).filter(models.Role.id == role_id).first()
    if not db_role:
        raise HTTPException(status_code=404, detail="Rôle introuvable")
    if db_role.is_builtin:
        raise HTTPException(status_code=400, detail="Les rôles intégrés ne peuvent pas être supprimés")
    user_count = db.query(models.User).filter(models.User.role == db_role.name).count()
    if user_count > 0:
        raise HTTPException(status_code=400, detail=f"Impossible de supprimer un rôle assigné à {user_count} utilisateur(s)")
    db.delete(db_role)
    db.commit()
    return {"deleted": 1, "name": db_role.name}


# ─── System Health ────────────────────────────────────────────────────────────

# Module-level snapshot for network I/O delta (updated on each call)
_net_snapshot: dict = {"time": 0.0, "bytes_recv": 0, "bytes_sent": 0}


@app.get("/api/system/health")
async def get_system_health(db: Session = Depends(get_db)):
    import psutil as _ps
    import time as _time
    import os as _os
    global _net_snapshot

    # ── Hardware ──────────────────────────────────────────────────────────────
    cpu_pct = _ps.cpu_percent(interval=None)
    mem     = _ps.virtual_memory()
    try:
        disk = _ps.disk_usage(".")
    except Exception:
        disk = _ps.disk_usage("/")

    # ── Network I/O delta (rate between consecutive calls) ────────────────────
    net     = _ps.net_io_counters()
    now     = _time.time()
    elapsed = now - _net_snapshot["time"]
    if elapsed > 0.5:
        recv_mbps = ((net.bytes_recv - _net_snapshot["bytes_recv"]) * 8) / (elapsed * 1_000_000)
        sent_mbps = ((net.bytes_sent - _net_snapshot["bytes_sent"]) * 8) / (elapsed * 1_000_000)
    else:
        recv_mbps = sent_mbps = 0.0
    _net_snapshot = {"time": now, "bytes_recv": net.bytes_recv, "bytes_sent": net.bytes_sent}

    # ── Temperature (Linux/Mac only; graceful skip on Windows) ────────────────
    temperature = None
    try:
        all_temps = _ps.sensors_temperatures()
        if all_temps:
            for key in ("coretemp", "cpu_thermal", "acpitz", "k10temp"):
                if key in all_temps and all_temps[key]:
                    temperature = round(all_temps[key][0].current, 1)
                    break
    except (AttributeError, NotImplementedError):
        pass

    # ── System uptime ─────────────────────────────────────────────────────────
    uptime_sec  = int(_time.time() - _ps.boot_time())
    uptime_days = uptime_sec // 86400
    uptime_hrs  = (uptime_sec % 86400) // 3600
    uptime_min  = (uptime_sec % 3600) // 60
    uptime_str  = f"{uptime_days}j {uptime_hrs}h {uptime_min}m"

    # ── Vision engine status ──────────────────────────────────────────────────
    v_eng        = vision_engine.get_vision_engine()
    engine_alive = bool(v_eng.running and v_eng.thread and v_eng.thread.is_alive())
    model_name   = _os.path.basename(v_eng.model_path) if v_eng.model_path else "—"

    # ── Database metrics ──────────────────────────────────────────────────────
    db_path    = "./cement_counter.db"
    db_size_mb = round(_os.path.getsize(db_path) / (1024 * 1024), 2) if _os.path.exists(db_path) else 0.0

    t0             = _time.perf_counter()
    total_logs     = db.query(models.DetectionLog).count()
    query_time_ms  = round((_time.perf_counter() - t0) * 1000, 2)
    total_sessions = db.query(models.Session).count()
    active_sess    = db.query(models.Session).filter(models.Session.status == "active").count()
    total_alerts   = db.query(models.AlertHistory).count()
    unread_alerts  = db.query(models.AlertHistory).filter(models.AlertHistory.is_read == False).count()

    # ── Services ──────────────────────────────────────────────────────────────
    services = [
        {
            "name":   "Moteur Vision YOLOv8",
            "status": "Opérationnel" if engine_alive else "Arrêté",
            "ok":     engine_alive,
            "uptime": uptime_str,
            "detail": model_name,
        },
        {
            "name":   "API FastAPI",
            "status": "Opérationnel",
            "ok":     True,
            "uptime": uptime_str,
            "detail": f"CPU {cpu_pct:.1f}%",
        },
        {
            "name":   "Base de Données SQLite",
            "status": "Opérationnel",
            "ok":     True,
            "uptime": uptime_str,
            "detail": f"{db_size_mb} MB · {query_time_ms} ms/req",
        },
        {
            "name":   "Flux Vidéo MJPEG",
            "status": "Actif" if engine_alive else "Inactif",
            "ok":     engine_alive,
            "uptime": uptime_str,
            "detail": "WebSocket /ws/video",
        },
    ]

    # ── Recent system events (alerts + sessions) ──────────────────────────────
    recent_alerts   = db.query(models.AlertHistory).order_by(
        models.AlertHistory.timestamp.desc()
    ).limit(5).all()
    recent_sessions = db.query(models.Session).order_by(
        models.Session.start_time.desc()
    ).limit(3).all()

    events: list = []
    for a in recent_alerts:
        lvl = "WARN" if a.alert_type == "warning" else ("ERROR" if a.alert_type == "critical" else "INFO")
        events.append({
            "timestamp": a.timestamp.isoformat(),
            "level":     lvl,
            "message":   f"[Alerte] {a.title or 'Alerte'} — {a.message}",
        })
    for s in recent_sessions:
        total = (s.total_count or 0) + (s.rejected_count or 0)
        events.append({
            "timestamp": s.start_time.isoformat(),
            "level":     "INFO",
            "message":   f"Session {s.id} — {total} sacs traités (statut: {s.status})",
        })
    events.sort(key=lambda e: e["timestamp"], reverse=True)
    events = events[:10]

    return {
        "status":    "online",
        "uptime":    uptime_str,
        "uptime_sec": uptime_sec,
        # Hardware
        "cpu":     round(cpu_pct, 1),
        "memory":  round(mem.percent, 1),
        "memory_total_gb": round(mem.total / (1024 ** 3), 1),
        "memory_used_gb":  round(mem.used  / (1024 ** 3), 1),
        "disk":    round(disk.percent, 1),
        "disk_total_gb": round(disk.total / (1024 ** 3), 1),
        "disk_used_gb":  round(disk.used  / (1024 ** 3), 1),
        "disk_free_gb":  round(disk.free  / (1024 ** 3), 1),
        "temperature": temperature,
        "net_recv_mbps": round(max(0.0, recv_mbps), 2),
        "net_sent_mbps": round(max(0.0, sent_mbps), 2),
        # Services
        "services": services,
        # DB stats
        "db": {
            "size_mb":         db_size_mb,
            "total_logs":      total_logs,
            "total_sessions":  total_sessions,
            "active_sessions": active_sess,
            "total_alerts":    total_alerts,
            "unread_alerts":   unread_alerts,
            "query_time_ms":   query_time_ms,
        },
        # Events log
        "events": events,
    }


# ─── Alerts ───────────────────────────────────────────────────────────────────

_ALERT_SETTING_KEYS = [
    "alert_sound_enabled", "alert_sound_volume",
    "alert_email_enabled", "alert_slack_enabled", "alert_supervisor_phone",
]


def _alert_type_from_rule(rule: models.AlertRule) -> str:
    if rule.type == "error_rate":
        return "critical"
    elif rule.type == "production_rate":
        return "warning"
    return "info"


def _alert_to_dict(alert: models.AlertHistory, rule: models.AlertRule | None = None) -> dict:
    atype = alert.alert_type or (_alert_type_from_rule(rule) if rule else "info")
    title = alert.title or (rule.name if rule else "Alerte")
    return {
        "id": alert.id,
        "rule_id": alert.rule_id,
        "timestamp": alert.timestamp.isoformat(),
        "message": alert.message,
        "title": title,
        "is_read": alert.is_read,
        "alert_type": atype,
    }


@app.get("/api/alerts/history")
async def get_alert_history(limit: int = 50, unread_only: bool = False, db: Session = Depends(get_db)):
    query = db.query(models.AlertHistory)
    if unread_only:
        query = query.filter(models.AlertHistory.is_read == False)
    alerts = query.order_by(models.AlertHistory.timestamp.desc()).limit(limit).all()
    rules_map = {r.id: r for r in db.query(models.AlertRule).all()}
    return [_alert_to_dict(a, rules_map.get(a.rule_id)) for a in alerts]


@app.get("/api/alerts/unread-count")
async def get_alert_unread_count(db: Session = Depends(get_db)):
    count = db.query(models.AlertHistory).filter(models.AlertHistory.is_read == False).count()
    return {"count": count}


@app.post("/api/alerts/history")
async def create_manual_alert(payload: schemas.AlertHistoryCreate, db: Session = Depends(get_db)):
    alert = models.AlertHistory(
        rule_id=None,
        message=payload.message,
        title=payload.title,
        alert_type=payload.alert_type,
        is_read=False,
    )
    db.add(alert)
    db.commit()
    db.refresh(alert)
    await manager.broadcast(json.dumps({"type": "ALERT_EVENT", "data": {
        "id": alert.id, "title": alert.title,
        "message": alert.message, "alert_type": alert.alert_type,
    }}))
    return _alert_to_dict(alert)


@app.patch("/api/alerts/history/{alert_id}/read")
async def mark_alert_read(alert_id: int, db: Session = Depends(get_db)):
    alert = db.query(models.AlertHistory).filter(models.AlertHistory.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="Alerte introuvable")
    alert.is_read = True
    db.commit()
    return {"ok": True}


@app.post("/api/alerts/history/read-all")
async def mark_all_alerts_read(db: Session = Depends(get_db)):
    db.query(models.AlertHistory).filter(
        models.AlertHistory.is_read == False
    ).update({"is_read": True})
    db.commit()
    return {"ok": True}


@app.delete("/api/alerts/history/{alert_id}")
async def delete_single_alert(alert_id: int, db: Session = Depends(get_db)):
    alert = db.query(models.AlertHistory).filter(models.AlertHistory.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="Alerte introuvable")
    db.delete(alert)
    db.commit()
    return {"ok": True}


@app.delete("/api/alerts/history")
async def clear_all_alerts(db: Session = Depends(get_db)):
    db.query(models.AlertHistory).delete()
    db.commit()
    return {"ok": True}


@app.get("/api/alerts/rules")
async def get_alert_rules(db: Session = Depends(get_db)):
    rules = db.query(models.AlertRule).all()
    return [{"id": r.id, "name": r.name, "type": r.type, "threshold": r.threshold, "is_active": r.is_active} for r in rules]


@app.put("/api/alerts/rules/{rule_id}")
async def update_alert_rule(rule_id: int, payload: schemas.AlertRuleUpdate, db: Session = Depends(get_db)):
    rule = db.query(models.AlertRule).filter(models.AlertRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Règle introuvable")
    if payload.name is not None:
        rule.name = payload.name
    if payload.threshold is not None:
        rule.threshold = payload.threshold
    if payload.is_active is not None:
        rule.is_active = payload.is_active
    db.commit()
    db.refresh(rule)
    return {"id": rule.id, "name": rule.name, "type": rule.type, "threshold": rule.threshold, "is_active": rule.is_active}


@app.get("/api/alerts/settings")
async def get_alert_settings(db: Session = Depends(get_db)):
    rows = db.query(models.SystemSetting).filter(
        models.SystemSetting.key.in_(_ALERT_SETTING_KEYS)
    ).all()
    s = {r.key: r.value for r in rows}
    return {
        "sound_enabled": s.get("alert_sound_enabled", "true").lower() == "true",
        "sound_volume": int(s.get("alert_sound_volume", "65")),
        "email_enabled": s.get("alert_email_enabled", "true").lower() == "true",
        "slack_enabled": s.get("alert_slack_enabled", "false").lower() == "true",
        "supervisor_phone": s.get("alert_supervisor_phone", "+33 6 12 34 56 78"),
    }


@app.put("/api/alerts/settings")
async def update_alert_settings(payload: schemas.AlertSettings, db: Session = Depends(get_db)):
    updates = {
        "alert_sound_enabled": str(payload.sound_enabled).lower(),
        "alert_sound_volume": str(payload.sound_volume),
        "alert_email_enabled": str(payload.email_enabled).lower(),
        "alert_slack_enabled": str(payload.slack_enabled).lower(),
        "alert_supervisor_phone": payload.supervisor_phone,
    }
    for key, value in updates.items():
        row = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
        if row:
            row.value = value
        else:
            db.add(models.SystemSetting(key=key, value=value))
    db.commit()
    return {"ok": True}


@app.post("/api/alerts/evaluate")
async def evaluate_alert_rules(db: Session = Depends(get_db)):
    """Evaluate active alert rules against recent production metrics and create alerts if triggered."""
    import datetime as _dt

    now = _dt.datetime.utcnow()
    window_5m = now - _dt.timedelta(minutes=5)
    window_10m = now - _dt.timedelta(minutes=10)

    rules = db.query(models.AlertRule).filter(models.AlertRule.is_active == True).all()
    triggered = []

    for rule in rules:
        # Anti-spam: skip if already alerted for this rule in last 10 minutes
        recent = db.query(models.AlertHistory).filter(
            models.AlertHistory.rule_id == rule.id,
            models.AlertHistory.timestamp >= window_10m,
        ).first()
        if recent:
            continue

        should_trigger = False
        message = ""
        alert_type = "warning"

        if rule.type == "production_rate":
            count = db.query(models.DetectionLog).filter(
                models.DetectionLog.timestamp >= window_5m,
                models.DetectionLog.status == "conforme",
            ).count()
            rate = count / 5.0
            if rate < rule.threshold:
                should_trigger = True
                alert_type = "warning"
                message = f"Cadence {rate:.1f} sacs/min sous le seuil de {rule.threshold:.1f} sacs/min."

        elif rule.type == "error_rate":
            total = db.query(models.DetectionLog).filter(
                models.DetectionLog.timestamp >= window_5m
            ).count()
            rejected = db.query(models.DetectionLog).filter(
                models.DetectionLog.timestamp >= window_5m,
                models.DetectionLog.status == "rejete",
            ).count()
            if total > 0:
                rate = (rejected / total) * 100
                if rate > rule.threshold:
                    should_trigger = True
                    alert_type = "critical"
                    message = f"Taux de rejet {rate:.1f}% dépasse le seuil de {rule.threshold:.1f}%."

        if should_trigger:
            alert = models.AlertHistory(
                rule_id=rule.id,
                message=message,
                title=rule.name,
                alert_type=alert_type,
                is_read=False,
            )
            db.add(alert)
            triggered.append({"rule": rule.name, "message": message})

    db.commit()
    for t in triggered:
        await manager.broadcast(json.dumps({"type": "ALERT_EVENT", "data": {
            "title": t["rule"], "message": t["message"], "alert_type": "critical",
        }}))
    return {"triggered": len(triggered), "alerts": triggered}


@app.post("/api/alerts/rules")
async def create_alert_rule(payload: schemas.AlertRuleBase, db: Session = Depends(get_db)):
    rule = models.AlertRule(
        name=payload.name,
        type=payload.type,
        threshold=payload.threshold,
        is_active=payload.is_active,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return {"id": rule.id, "name": rule.name, "type": rule.type, "threshold": rule.threshold, "is_active": rule.is_active}


@app.delete("/api/alerts/rules/{rule_id}")
async def delete_alert_rule(rule_id: int, db: Session = Depends(get_db)):
    rule = db.query(models.AlertRule).filter(models.AlertRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Règle introuvable")
    db.delete(rule)
    db.commit()
    return {"ok": True}


# ─── System General Settings ──────────────────────────────────────────────────

_GENERAL_SETTING_KEYS = [
    "site_name", "site_location", "site_timezone", "site_language",
    "notify_low_production", "notify_weekly_reports",
    "log_level", "log_retention_days",
    "cache_max_gb", "cache_auto_cleanup",
]

_GENERAL_DEFAULTS: dict = {
    "site_name": "Cimenterie Centrale - Ligne A",
    "site_location": "Zone Industrielle Nord, Secteur 4",
    "site_timezone": "utc1",
    "site_language": "fr",
    "notify_low_production": "true",
    "notify_weekly_reports": "true",
    "log_level": "info",
    "log_retention_days": "30",
    "cache_max_gb": "10",
    "cache_auto_cleanup": "true",
}

_SECURITY_SETTING_KEYS = [
    "jwt_expire_minutes", "max_login_attempts",
    "session_timeout_minutes", "require_2fa_admin", "force_https",
]

_SECURITY_DEFAULTS: dict = {
    "jwt_expire_minutes": "30",
    "max_login_attempts": "5",
    "session_timeout_minutes": "480",
    "require_2fa_admin": "false",
    "force_https": "false",
}


@app.get("/api/system/general-settings")
async def get_general_settings(db: Session = Depends(get_db)):
    rows = db.query(models.SystemSetting).filter(
        models.SystemSetting.key.in_(_GENERAL_SETTING_KEYS)
    ).all()
    s = {r.key: r.value for r in rows}
    return {k: s.get(k, _GENERAL_DEFAULTS[k]) for k in _GENERAL_SETTING_KEYS}


@app.put("/api/system/general-settings")
async def update_general_settings(payload: dict, db: Session = Depends(get_db)):
    for key, value in payload.items():
        if key not in _GENERAL_SETTING_KEYS:
            continue
        row = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
        if row:
            row.value = str(value)
        else:
            db.add(models.SystemSetting(key=key, value=str(value)))
    db.commit()
    return {"ok": True}


@app.get("/api/system/security-settings")
async def get_security_settings(db: Session = Depends(get_db)):
    rows = db.query(models.SystemSetting).filter(
        models.SystemSetting.key.in_(_SECURITY_SETTING_KEYS)
    ).all()
    s = {r.key: r.value for r in rows}
    return {k: s.get(k, _SECURITY_DEFAULTS[k]) for k in _SECURITY_SETTING_KEYS}


@app.put("/api/system/security-settings")
async def update_security_settings(payload: dict, db: Session = Depends(get_db)):
    for key, value in payload.items():
        if key not in _SECURITY_SETTING_KEYS:
            continue
        row = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
        if row:
            row.value = str(value)
        else:
            db.add(models.SystemSetting(key=key, value=str(value)))
    db.commit()
    return {"ok": True}


@app.get("/api/system/export-config")
async def export_config(db: Session = Depends(get_db)):
    """Export all system settings + cameras as a downloadable JSON file."""
    import json as _j
    rows = db.query(models.SystemSetting).all()
    cameras = db.query(models.Camera).all()
    data = {
        "version": "1.0",
        "exported_at": datetime.utcnow().isoformat(),
        "settings": {r.key: r.value for r in rows},
        "cameras": [
            {
                "name": c.name, "source_type": c.source_type,
                "url": c.url, "resolution": c.resolution,
                "fps": c.fps, "notes": c.notes,
            }
            for c in cameras
        ],
    }
    content = _j.dumps(data, indent=2, ensure_ascii=False)
    fname = f"config_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
    return Response(
        content=content,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.post("/api/system/import-config")
async def import_config(payload: dict, db: Session = Depends(get_db)):
    """Restore system settings from an exported config JSON."""
    settings = payload.get("settings", {})
    if not isinstance(settings, dict):
        raise HTTPException(status_code=400, detail="Format invalide : 'settings' doit être un objet JSON")
    restored = 0
    for key, value in settings.items():
        row = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
        if row:
            row.value = str(value)
        else:
            db.add(models.SystemSetting(key=key, value=str(value)))
        restored += 1
    db.commit()
    return {"ok": True, "restored_keys": restored}


@app.get("/api/system/db-backup")
async def download_db_backup():
    """Download the SQLite database file as a binary attachment."""
    import os as _os
    from fastapi.responses import FileResponse as _FR
    db_path = _os.path.abspath(
        _os.path.join(_os.path.dirname(__file__), "..", "production.db")
    )
    if not _os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="Fichier de base de données introuvable")
    fname = f"production_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.db"
    return _FR(path=db_path, media_type="application/octet-stream", filename=fname)


# ─── Profile self-service ──────────────────────────────────────────────────────
class _SelfPasswordChange(BaseModel):
    current_password: str
    new_password: str

@app.patch("/api/users/me/password")
async def change_own_password(
    payload: _SelfPasswordChange,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    if not auth.verify_password(payload.current_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Mot de passe actuel incorrect")
    if len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="Le nouveau mot de passe doit contenir au moins 6 caractères")
    current_user.hashed_password = auth.get_password_hash(payload.new_password)
    db.add(models.UserActivity(user_id=current_user.id, username=current_user.username, action="password_changed"))
    db.commit()
    return {"ok": True}


# ─── Audit Trail ──────────────────────────────────────────────────────────────
@app.get("/api/audit/")
async def get_audit_trail(
    page: int = 1,
    page_size: int = 20,
    action: str | None = None,
    username: str | None = None,
    db: Session = Depends(get_db),
):
    page = max(1, page)
    page_size = min(max(page_size, 1), 100)
    q = db.query(models.UserActivity).order_by(models.UserActivity.timestamp.desc())
    if action:
        q = q.filter(models.UserActivity.action == action)
    if username:
        q = q.filter(models.UserActivity.username.ilike(f"%{username}%"))
    total = q.count()
    rows = q.offset((page - 1) * page_size).limit(page_size).all()
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": r.id,
                "username": r.username,
                "timestamp": r.timestamp.isoformat(),
                "action": r.action,
                "ip_address": r.ip_address,
                "user_agent": r.user_agent,
            }
            for r in rows
        ],
    }


# ─── API Keys ─────────────────────────────────────────────────────────────────
import secrets as _secrets
import hashlib as _hashlib

def _hash_key(raw_key: str) -> str:
    return _hashlib.sha256(raw_key.encode()).hexdigest()

class _ApiKeyCreate(BaseModel):
    name: str
    scope: str = "read"

@app.get("/api/apikeys/")
async def list_api_keys(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    keys = db.query(models.ApiKey).filter(models.ApiKey.is_active == True).order_by(models.ApiKey.created_at.desc()).all()
    return [
        {
            "id": k.id,
            "name": k.name,
            "key_prefix": k.key_prefix,
            "scope": k.scope,
            "user_id": k.user_id,
            "created_at": k.created_at.isoformat() if k.created_at else None,
            "last_used_at": k.last_used_at.isoformat() if k.last_used_at else None,
        }
        for k in keys
    ]

@app.post("/api/apikeys/", status_code=201)
async def create_api_key(
    payload: _ApiKeyCreate,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    raw_key = "cmt_" + _secrets.token_hex(24)
    key_prefix = raw_key[:12] + "..."
    new_key = models.ApiKey(
        name=payload.name,
        key_prefix=key_prefix,
        key_hash=_hash_key(raw_key),
        scope=payload.scope,
        user_id=current_user.id,
        is_active=True,
    )
    db.add(new_key)
    db.commit()
    db.refresh(new_key)
    return {
        "id": new_key.id,
        "name": new_key.name,
        "key_prefix": new_key.key_prefix,
        "scope": new_key.scope,
        "raw_key": raw_key,  # affiché une seule fois
        "created_at": new_key.created_at.isoformat() if new_key.created_at else None,
    }

@app.delete("/api/apikeys/{key_id}")
async def revoke_api_key(
    key_id: int,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    key = db.query(models.ApiKey).filter(models.ApiKey.id == key_id).first()
    if not key:
        raise HTTPException(status_code=404, detail="Clé introuvable")
    key.is_active = False
    db.commit()
    return {"ok": True}


# ─── Integration Settings ─────────────────────────────────────────────────────
_INTEGRATION_KEYS = [
    "webhook_enabled", "webhook_url", "webhook_secret",
    "smtp_host", "smtp_port", "smtp_user", "smtp_password", "smtp_from",
    "slack_webhook_url", "slack_enabled",
    "teams_webhook_url", "teams_enabled",
]
_INTEGRATION_DEFAULTS: dict = {
    "webhook_enabled": "false", "webhook_url": "", "webhook_secret": "",
    "smtp_host": "", "smtp_port": "587", "smtp_user": "", "smtp_password": "", "smtp_from": "",
    "slack_webhook_url": "", "slack_enabled": "false",
    "teams_webhook_url": "", "teams_enabled": "false",
}

@app.get("/api/system/integration-settings")
async def get_integration_settings(db: Session = Depends(get_db)):
    rows = db.query(models.SystemSetting).filter(models.SystemSetting.key.in_(_INTEGRATION_KEYS)).all()
    s = {r.key: r.value for r in rows}
    return {k: s.get(k, _INTEGRATION_DEFAULTS[k]) for k in _INTEGRATION_KEYS}

@app.put("/api/system/integration-settings")
async def update_integration_settings(payload: dict, db: Session = Depends(get_db)):
    for key, value in payload.items():
        if key not in _INTEGRATION_KEYS:
            continue
        row = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
        if row:
            row.value = str(value)
        else:
            db.add(models.SystemSetting(key=key, value=str(value)))
    db.commit()
    return {"ok": True}

def _get_smtp_config(db) -> dict:
    """Read SMTP settings from DB and return as dict."""
    keys = ["smtp_host", "smtp_port", "smtp_user", "smtp_password", "smtp_from"]
    rows = db.query(models.SystemSetting).filter(models.SystemSetting.key.in_(keys)).all()
    s = {r.key: r.value for r in rows}
    return {
        "host":     s.get("smtp_host", ""),
        "port":     int(s.get("smtp_port", "587")),
        "user":     s.get("smtp_user", ""),
        "password": s.get("smtp_password", ""),
        "from":     s.get("smtp_from", ""),
    }


def _send_email(smtp_cfg: dict, to: str, subject: str, body: str, attachment_path: str | None = None):
    """Send an email via SMTP (TLS/STARTTLS). Raises on failure."""
    import smtplib
    import ssl as _ssl
    from email.message import EmailMessage
    from pathlib import Path as _P

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"]    = smtp_cfg["from"] or smtp_cfg["user"]
    msg["To"]      = to
    msg.set_content(body)

    if attachment_path:
        p = _P(attachment_path)
        if p.exists():
            ctype = "application/octet-stream"
            if p.suffix == ".csv":   ctype = "text/csv"
            elif p.suffix == ".xlsx": ctype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            elif p.suffix == ".pdf":  ctype = "application/pdf"
            elif p.suffix == ".json": ctype = "application/json"
            maintype, subtype = ctype.split("/", 1)
            msg.add_attachment(p.read_bytes(), maintype=maintype, subtype=subtype, filename=p.name)

    ctx = _ssl.create_default_context()
    with smtplib.SMTP(smtp_cfg["host"], smtp_cfg["port"], timeout=10) as server:
        server.ehlo()
        server.starttls(context=ctx)
        server.login(smtp_cfg["user"], smtp_cfg["password"])
        server.send_message(msg)


@app.post("/api/system/test-smtp")
async def test_smtp(db: Session = Depends(get_db)):
    """Send a test email using the configured SMTP settings."""
    cfg = _get_smtp_config(db)
    if not cfg["host"]:
        raise HTTPException(status_code=400, detail="Serveur SMTP non configuré (smtp_host manquant).")
    if not cfg["user"]:
        raise HTTPException(status_code=400, detail="Utilisateur SMTP non configuré (smtp_user manquant).")
    to = cfg["from"] or cfg["user"]
    try:
        _send_email(
            cfg,
            to=to,
            subject="[CimentMonitor] Test de connexion SMTP",
            body=(
                "Bonjour,\n\n"
                "Ceci est un email de test envoyé par CimentMonitor Pro.\n"
                "La configuration SMTP est correctement opérationnelle.\n\n"
                "— CimentMonitor Pro"
            ),
        )
        return {"ok": True, "message": f"Email de test envoyé à {to}"}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Échec envoi SMTP : {str(e)}")


@app.post("/api/system/test-webhook")
async def test_webhook(db: Session = Depends(get_db)):
    import httpx as _httpx
    rows = db.query(models.SystemSetting).filter(models.SystemSetting.key.in_(["webhook_url", "webhook_secret"])).all()
    s = {r.key: r.value for r in rows}
    url = s.get("webhook_url", "")
    secret = s.get("webhook_secret", "")
    if not url:
        raise HTTPException(status_code=400, detail="URL du webhook non configurée")
    try:
        headers = {"Content-Type": "application/json"}
        if secret:
            headers["X-Webhook-Secret"] = secret
        async with _httpx.AsyncClient(timeout=5.0) as client:
            r = await client.post(url, json={"event": "test", "source": "CimentMonitorPro"}, headers=headers)
        return {"ok": True, "status_code": r.status_code}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Échec: {str(e)}")


# ─── Analytics / OEE ─────────────────────────────────────────────────────────
@app.get("/api/analytics/oee")
async def get_oee_analytics(hours: int = 24, db: Session = Depends(get_db)):
    """OEE / TRS analytics computed from real DetectionLog & Session data."""
    import math as _math
    from datetime import datetime as _dt, timedelta as _td

    hours = max(1, min(hours, 720))
    now = _dt.utcnow()
    start = now - _td(hours=hours)
    prev_start = start - _td(hours=hours)
    planned_seconds = hours * 3600
    TARGET_RATE = 1100  # bags/hour theoretical max

    def _to_dt(ts):
        if isinstance(ts, str):
            try:
                return _dt.fromisoformat(ts)
            except Exception:
                return None
        return ts

    # ── Current period ──────────────────────────────────────────────────────
    logs = (
        db.query(models.DetectionLog)
        .filter(models.DetectionLog.timestamp >= start)
        .order_by(models.DetectionLog.timestamp.asc())
        .all()
    )
    conforming = [l for l in logs if l.status == "conforme"]
    rejected   = [l for l in logs if l.status == "rejete"]
    total_bags     = len(conforming)
    rejected_count = len(rejected)
    total_inspected = total_bags + rejected_count

    quality = round((total_bags / total_inspected * 100) if total_inspected > 0 else 100.0, 1)

    # Sessions overlapping the period
    sessions = db.query(models.Session).filter(
        models.Session.start_time >= start
    ).all()
    active_sess = db.query(models.Session).filter(models.Session.status == "active").first()
    if active_sess and active_sess not in sessions:
        sessions.append(active_sess)

    total_session_seconds = 0.0
    for sess in sessions:
        s_start = _to_dt(sess.start_time)
        s_end   = _to_dt(sess.end_time) if sess.end_time else now
        if s_start is None:
            continue
        s_start = max(s_start, start)
        s_end   = min(s_end,   now)
        dur = (s_end - s_start).total_seconds()
        if dur > 0:
            total_session_seconds += dur

    availability = round(
        min(100.0, total_session_seconds / planned_seconds * 100) if planned_seconds > 0 else 0.0, 1
    )

    actual_rate = total_bags / (total_session_seconds / 3600) if total_session_seconds > 0 else 0.0
    performance = round(min(100.0, actual_rate / TARGET_RATE * 100) if TARGET_RATE > 0 else 0.0, 1)

    oee = round((availability * performance * quality) / 10000.0, 1)

    # ── Previous period OEE (for delta) ────────────────────────────────────
    prev_logs = db.query(models.DetectionLog).filter(
        models.DetectionLog.timestamp >= prev_start,
        models.DetectionLog.timestamp < start,
    ).all()
    prev_conforming = [l for l in prev_logs if l.status == "conforme"]
    prev_total      = len(prev_conforming)
    prev_inspected  = prev_total + len([l for l in prev_logs if l.status == "rejete"])

    prev_sessions = db.query(models.Session).filter(
        models.Session.start_time >= prev_start,
        models.Session.start_time <  start,
    ).all()
    prev_sess_secs = 0.0
    for sess in prev_sessions:
        s_start = _to_dt(sess.start_time)
        s_end   = _to_dt(sess.end_time) if sess.end_time else start
        if s_start is None:
            continue
        s_start = max(s_start, prev_start)
        s_end   = min(s_end,   start)
        dur = (s_end - s_start).total_seconds()
        if dur > 0:
            prev_sess_secs += dur

    prev_quality      = (prev_total / prev_inspected * 100) if prev_inspected > 0 else 100.0
    prev_availability = min(100.0, prev_sess_secs / planned_seconds * 100) if planned_seconds > 0 else 0.0
    prev_actual_rate  = prev_total / (prev_sess_secs / 3600) if prev_sess_secs > 0 else 0.0
    prev_performance  = min(100.0, prev_actual_rate / TARGET_RATE * 100) if TARGET_RATE > 0 else 0.0
    prev_oee          = (prev_availability * prev_performance * prev_quality) / 10000.0
    oee_delta = round(oee - prev_oee, 1)

    # ── Hourly / bucketed chart data ────────────────────────────────────────
    if hours <= 48:
        bucket_hours = 1
    elif hours <= 168:
        bucket_hours = 4
    else:
        bucket_hours = 24

    num_buckets = hours // bucket_hours
    hourly_data = []
    for i in range(num_buckets - 1, -1, -1):
        b_start = now - _td(hours=(i + 1) * bucket_hours)
        b_end   = now - _td(hours=i * bucket_hours)
        if bucket_hours >= 24:
            label = b_start.strftime("%d/%m")
        elif bucket_hours > 1:
            label = b_start.strftime("%d/%m %H:%M")
        else:
            label = b_start.strftime("%H:%M")

        b_count = sum(
            1 for l in conforming
            if _to_dt(l.timestamp) and b_start <= _to_dt(l.timestamp) < b_end
        )
        hourly_data.append({"name": label, "real": b_count, "target": TARGET_RATE, "forecast": 0})

    # Rolling 3-bucket forecast
    for i, h in enumerate(hourly_data):
        prev_vals = [hourly_data[j]["real"] for j in range(max(0, i - 3), i) if hourly_data[j]["real"] > 0]
        h["forecast"] = round(sum(prev_vals) / len(prev_vals)) if prev_vals else h["real"]

    # ── Downtime distribution ───────────────────────────────────────────────
    all_ts = sorted(ts for l in logs if (ts := _to_dt(l.timestamp)) is not None)

    production_secs    = 0.0
    micro_stops_secs   = 0.0
    technical_fail_secs = 0.0
    for i in range(1, len(all_ts)):
        delta = (all_ts[i] - all_ts[i - 1]).total_seconds()
        if delta < 30:
            production_secs    += delta
        elif delta < 120:
            micro_stops_secs   += delta
        else:
            technical_fail_secs += delta

    if planned_seconds > 0:
        prod_pct    = round(production_secs     / planned_seconds * 100, 1)
        micro_pct   = round(micro_stops_secs    / planned_seconds * 100, 1)
        failure_pct = round(technical_fail_secs / planned_seconds * 100, 1)
        idle_pct    = max(0.0, round(100 - prod_pct - micro_pct - failure_pct, 1))
    else:
        prod_pct = micro_pct = failure_pct = idle_pct = 0.0

    downtime_data = [
        {"name": "Production",      "value": prod_pct,    "color": "#22c55e"},
        {"name": "Micro-arrêts",    "value": micro_pct,   "color": "#eab308"},
        {"name": "Panne Technique", "value": failure_pct, "color": "#ef4444"},
        {"name": "Inactivité",      "value": idle_pct,    "color": "#6366f1"},
    ]
    total_stops_h = round((planned_seconds - production_secs) / 3600, 1)

    # ── Recommendations ─────────────────────────────────────────────────────
    intervals = []
    for i in range(1, len(all_ts)):
        d = (all_ts[i] - all_ts[i - 1]).total_seconds()
        if 0 < d < 120:
            intervals.append(d)

    recommendations = []
    if intervals:
        avg_iv = sum(intervals) / len(intervals)
        if len(intervals) >= 2:
            variance = sum((x - avg_iv) ** 2 for x in intervals) / len(intervals)
            cv = _math.sqrt(variance) / avg_iv if avg_iv > 0 else 0
            if cv > 0.3:
                recommendations.append({
                    "type": "speed", "color": "orange",
                    "title": "Vitesse Convoyeur",
                    "text": (
                        f"Coefficient de variation des intervalles : {cv*100:.0f}%. "
                        f"Stabiliser la cadence pourrait améliorer la consistance "
                        f"de {min(15, int(cv * 30))}%."
                    ),
                })

    if total_inspected > 0 and rejected_count / total_inspected > 0.03:
        recommendations.append({
            "type": "quality", "color": "blue",
            "title": "Taux de Rejet",
            "text": (
                f"{rejected_count} sacs rejetés "
                f"({rejected_count / total_inspected * 100:.1f}% du total). "
                f"Vérifier le positionnement caméra et le seuil de confiance du modèle."
            ),
        })

    if availability < 80.0 and total_bags > 0:
        recommendations.append({
            "type": "maintenance", "color": "red",
            "title": "Disponibilité Faible",
            "text": (
                f"Disponibilité à {availability}% sur la période. "
                f"Augmenter la durée des sessions de production actives."
            ),
        })

    defaults = [
        {"type": "speed", "color": "orange", "title": "Vitesse Convoyeur",
         "text": "Débit stable sur la période analysée. Aucun ajustement de vitesse nécessaire."},
        {"type": "maintenance", "color": "blue", "title": "Maintenance Prédictive",
         "text": "Aucune anomalie de cadence détectée. Prochain entretien selon le calendrier prévu."},
        {"type": "quality", "color": "green", "title": "Qualité de Donnée",
         "text": f"Taux de conformité à {quality}%. Les conditions de détection sont optimales."},
    ]
    while len(recommendations) < 3:
        recommendations.append(defaults[len(recommendations) % len(defaults)])

    return {
        "oee":              oee,
        "oeeDelta":         oee_delta,
        "availability":     availability,
        "performance":      performance,
        "quality":          quality,
        "totalBags":        total_bags,
        "rejectedBags":     rejected_count,
        "targetRatePerHour": TARGET_RATE,
        "actualRatePerHour": round(actual_rate, 1),
        "sessionHours":     round(total_session_seconds / 3600, 1),
        "hourlyData":       hourly_data,
        "downtimeData":     downtime_data,
        "totalStopsHours":  total_stops_h,
        "recommendations":  recommendations,
        "periodHours":      hours,
    }


# ─── Quality ──────────────────────────────────────────────────────────────────
@app.get("/api/quality/summary", response_model=schemas.QualityDashboardResponse)
async def get_quality_summary(db: Session = Depends(get_db)):
    logs = db.query(models.DetectionLog).all()
    total = len(logs)
    rejected = len([l for l in logs if l.status == "rejete"])
    rejection_rate = (rejected / total * 100) if total > 0 else 0

    avg_logo = sum((l.logo_score or 0) for l in logs) / total if total else 0
    avg_color = sum((l.color_score or 0) for l in logs) / total if total else 0
    avg_detect = sum((l.detection_score or 0) for l in logs) / total if total else 0

    bins = {"0-20%": 0, "20-40%": 0, "40-60%": 0, "60-80%": 0, "80-100%": 0}
    for l in logs:
        v = (l.detection_score or 0) * 100
        if v < 20:
            bins["0-20%"] += 1
        elif v < 40:
            bins["20-40%"] += 1
        elif v < 60:
            bins["40-60%"] += 1
        elif v < 80:
            bins["60-80%"] += 1
        else:
            bins["80-100%"] += 1

    logo_conforme = len([l for l in logs if (l.logo_score or 0) >= 0.8])
    logo_flou = len([l for l in logs if 0.5 <= (l.logo_score or 0) < 0.8])
    logo_absent = len([l for l in logs if (l.logo_score or 0) < 0.5])
    reviews_count = db.query(models.QualityReview).count()
    recent_errors = db.query(models.DetectionLog).filter(
        models.DetectionLog.status == "rejete",
        models.DetectionLog.timestamp >= (datetime.utcnow() - timedelta(hours=24))
    ).count()

    return {
        "totalInspected": total,
        "rejectedCount": rejected,
        "rejectionRate": rejection_rate,
        "avgLogoScore": avg_logo,
        "avgColorScore": avg_color,
        "avgDetectionScore": avg_detect,
        "confidenceDistribution": [{"range": k, "count": v} for k, v in bins.items()],
        "logoDistribution": [
            {"name": "Logo Conforme", "value": logo_conforme, "color": "#f97316"},
            {"name": "Logo Flou", "value": logo_flou, "color": "#eab308"},
            {"name": "Sans Logo", "value": logo_absent, "color": "#ef4444"},
        ],
        "recentErrors": recent_errors,
        "reviewedCorrections": reviews_count,
    }


@app.get("/api/quality/manual-verification", response_model=schemas.ManualVerificationResponse)
async def get_manual_verification_queue(
    page: int = 1,
    page_size: int = 20,
    search: str | None = None,
    db: Session = Depends(get_db),
):
    page = max(page, 1)
    page_size = min(max(page_size, 1), 100)

    reviewed_ids = [r[0] for r in db.query(models.QualityReview.log_id).all()]

    query = db.query(models.DetectionLog).filter(
        (models.DetectionLog.status == "rejete") |
        (models.DetectionLog.detection_score < 0.6)
    )
    if reviewed_ids:
        query = query.filter(~models.DetectionLog.id.in_(reviewed_ids))
    if search:
        query = query.filter(models.DetectionLog.identifier.ilike(f"%{search}%"))

    total = query.count()
    items = query.order_by(models.DetectionLog.timestamp.desc()).offset((page - 1) * page_size).limit(page_size).all()

    mapped = []
    for l in items:
        reason = "Non conforme" if l.status == "rejete" else "Confiance faible"
        mapped.append({
            "id": l.id,
            "timestamp": l.timestamp,
            "session_id": l.session_id,
            "identifier": l.identifier,
            "detection_score": l.detection_score,
            "logo_score": l.logo_score,
            "color_score": l.color_score,
            "interval": l.interval,
            "capture_url": l.capture_url,
            "status": l.status,
            "reason": reason,
            "reviewed": False,
        })
    return {"items": mapped, "total": total}


@app.get("/api/quality/reviews", response_model=list[schemas.QualityReview])
async def get_quality_reviews(limit: int = 50, db: Session = Depends(get_db)):
    return db.query(models.QualityReview).order_by(models.QualityReview.created_at.desc()).limit(limit).all()


@app.get("/api/quality/anomalies", response_model=schemas.QualityAnomalyResponse)
async def get_quality_anomalies(limit: int = 50, db: Session = Depends(get_db)):
    logs = db.query(models.DetectionLog).order_by(models.DetectionLog.timestamp.desc()).limit(limit).all()
    items = []
    for l in logs:
        is_low_conf = (l.detection_score or 0) < 0.6
        is_reject = l.status == "rejete"
        if not is_low_conf and not is_reject:
            continue
        if getattr(l, "is_resolved", False):
            continue
        severity = "high" if (l.detection_score or 0) < 0.5 or is_reject else "medium"
        items.append({
            "id": f"AN-{l.id}",
            "type": "Sac rejeté" if is_reject else "Confiance faible",
            "time": l.timestamp.strftime("%H:%M:%S"),
            "severity": severity,
            "description": f"Score détection {l.detection_score:.2f}, logo {l.logo_score:.2f}, couleur {l.color_score:.2f}",
            "thumbnail": f"/static/captures/{l.capture_url.split('/')[-1]}" if l.capture_url else None,
            "status": "pending",
        })
    return {"items": items, "total": len(items)}


@app.patch("/api/quality/anomalies/{log_id}/resolve")
async def resolve_anomaly(log_id: int, db: Session = Depends(get_db)):
    log = db.query(models.DetectionLog).filter(models.DetectionLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Anomalie introuvable.")
    log.is_resolved = True
    db.commit()
    return {"success": True, "id": log_id}


@app.post("/api/quality/anomalies/resolve-all")
async def resolve_all_anomalies(db: Session = Depends(get_db)):
    logs = db.query(models.DetectionLog).filter(
        (models.DetectionLog.status == "rejete") |
        (models.DetectionLog.detection_score < 0.6)
    ).all()
    count = 0
    for log in logs:
        if not getattr(log, "is_resolved", False):
            log.is_resolved = True
            count += 1
    db.commit()
    return {"success": True, "resolved_count": count}


@app.patch("/api/logs/{log_id}", response_model=schemas.DetectionLog)
async def review_log(log_id: int, payload: schemas.UpdateLogRequest, db: Session = Depends(get_db)):
    log = db.query(models.DetectionLog).filter(models.DetectionLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")

    original_status = log.status

    action = payload.action.lower()
    if action in ["validate", "valider", "correct"]:
        log.status = payload.target_status or "conforme"
    elif action in ["reject", "rejeter"]:
        log.status = payload.target_status or "rejete"
    elif action in ["ignore", "ignorer"]:
        pass

    if payload.corrected_identifier:
        log.identifier = payload.corrected_identifier

    review = models.QualityReview(
        log_id=log.id,
        action=payload.action,
        target_status=log.status,
        notes=payload.notes,
        reviewer=payload.reviewer,
    )
    db.add(review)

    if original_status != log.status:
        _recompute_session_counters(db, log.session_id)

    db.commit()
    db.refresh(log)
    return log


# ─── Camera Configuration ────────────────────────────────────────────────────
@app.get("/api/config/camera", response_model=schemas.CameraConfig)
async def get_camera_config(db: Session = Depends(get_db)):
    settings = db.query(models.SystemSetting).filter(
        models.SystemSetting.key.in_([
            "camera_source_type", "camera_url", "camera_resolution",
            "camera_fps", "camera_brightness", "camera_contrast", "camera_autofocus"
        ])
    ).all()
    config = {s.key: s.value for s in settings}
    return schemas.CameraConfig(
        source_type=config.get("camera_source_type", "webcam"),
        url=config.get("camera_url", "0"),
        resolution=config.get("camera_resolution", "720p"),
        fps=int(config.get("camera_fps", "30")),
        brightness=int(config.get("camera_brightness", "50")),
        contrast=int(config.get("camera_contrast", "65")),
        autofocus=config.get("camera_autofocus", "true").lower() == "true",
    )


@app.put("/api/config/camera", response_model=schemas.CameraConfig)
async def update_camera_config(config: schemas.CameraConfig, db: Session = Depends(get_db)):
    mapping = {
        "camera_source_type": config.source_type,
        "camera_url": config.url,
        "camera_resolution": config.resolution,
        "camera_fps": str(config.fps),
        "camera_brightness": str(config.brightness),
        "camera_contrast": str(config.contrast),
        "camera_autofocus": str(config.autofocus).lower(),
    }
    for key, value in mapping.items():
        setting = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
        if setting:
            setting.value = value
        else:
            db.add(models.SystemSetting(key=key, value=value))
    db.commit()

    # Determine new video source and apply camera settings in runtime
    v_engine = vision_engine.get_vision_engine()
    if config.source_type == "webcam":
        new_source = int(config.url) if config.url.isdigit() else 0
    else:
        new_source = config.url

    source_changed = v_engine.video_source != new_source
    if source_changed:
        v_engine.video_source = new_source

    # Apply settings (hardware + software post-processing parameters)
    v_engine.apply_camera_settings(
        resolution=config.resolution,
        fps=config.fps,
        brightness=config.brightness,
        contrast=config.contrast,
        autofocus=config.autofocus,
    )

    # Restart only when source changes (avoid unnecessary RTSP reconnect storms)
    if v_engine.running and source_changed:
        if v_engine.stop():
            v_engine.start()

    return config


@app.post("/api/config/camera/test", response_model=schemas.CameraTestResult)
async def test_camera_connection(config: schemas.CameraConfig):
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(None, _test_camera_sync, config)
    return result


@app.get("/api/config/runtime")
async def get_runtime_config(db: Session = Depends(get_db)):
    v_engine = vision_engine.get_vision_engine()
    settings = db.query(models.SystemSetting).filter(
        models.SystemSetting.key.in_(["camera_url", "detection_model_path"])
    ).all()
    cfg = {s.key: s.value for s in settings}
    runtime = v_engine.get_runtime_info()
    return {
        "camera_name": runtime.get("camera_name", f"CAM_{cfg.get('camera_url', '0')}"),
        "model": cfg.get("detection_model_path", runtime.get("model", "models/best_V5.pt")),
        "capture_fps": runtime.get("capture_fps", 0),
        "line": runtime.get("line", {}),
    }


def _test_camera_sync(config: schemas.CameraConfig) -> dict:
    """Test camera connection synchronously (runs in thread pool).

    If the tested source matches the running vision engine source,
    temporarily stop the engine to avoid webcam access conflicts on Windows.
    """
    import os as _os
    import sys as _sys

    if config.source_type == "webcam":
        source = int(config.url) if config.url.isdigit() else 0
    else:
        source = config.url

    # Check if we need to pause the vision engine (same webcam source conflict only)
    v_engine = vision_engine.get_vision_engine()
    engine_was_running = False
    same_source = v_engine.running and v_engine.video_source == source
    is_webcam_source = isinstance(source, int) or (isinstance(source, str) and source.isdigit())
    if same_source and is_webcam_source:
        print(f"INFO: Pause du moteur de vision pour test caméra (source webcam: {source})")
        engine_was_running = v_engine.stop()

    cap = None
    try:
        # Use FFMPEG backend for RTSP/HTTP
        if isinstance(source, str) and source.startswith("rtsp://"):
            _os.environ["OPENCV_FFMPEG_CAPTURE_OPTIONS"] = "rtsp_transport;tcp|analyzeduration;5000000|probesize;5000000"
            cap = cv2.VideoCapture(source, cv2.CAP_FFMPEG)
        elif isinstance(source, str) and (source.startswith("http://") or source.startswith("https://")):
            cap = cv2.VideoCapture(source, cv2.CAP_FFMPEG)
        elif isinstance(source, int) or (isinstance(source, str) and source.isdigit()):
            # Webcam: use DirectShow on Windows
            idx = int(source) if isinstance(source, str) else source
            if _sys.platform == "win32":
                cap = cv2.VideoCapture(idx, cv2.CAP_DSHOW)
            else:
                cap = cv2.VideoCapture(idx)
        else:
            # Video file
            cap = cv2.VideoCapture(source)

        if not cap.isOpened():
            return {"success": False, "message": f"Impossible d'ouvrir la source: {source}"}

        ret, frame = cap.read()
        if not ret or frame is None:
            return {"success": False, "message": "Source ouverte mais aucune image reçue."}

        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        fps = cap.get(cv2.CAP_PROP_FPS)

        return {
            "success": True,
            "message": f"Connexion réussie — {width}x{height} @ {fps:.1f} FPS",
            "resolution_detected": f"{width}x{height}",
            "fps_detected": round(fps, 1),
        }
    except Exception as e:
        return {"success": False, "message": f"Erreur: {str(e)}"}
    finally:
        if cap is not None:
            cap.release()
        # Restart vision engine if we paused it
        if engine_was_running:
            print("INFO: Redémarrage du moteur de vision après test caméra")
            v_engine.start()


# ─── Available Models ────────────────────────────────────────────────────────

@app.get("/api/models/list")
async def list_available_models():
    """Liste les fichiers .pt disponibles dans le dossier models/."""
    import os as _os
    models_dir = "models"
    v_engine = vision_engine.get_vision_engine()
    active_path = v_engine.model_path
    if not _os.path.isdir(models_dir):
        return {"models": [], "active_model": active_path}
    result = []
    for fname in sorted(_os.listdir(models_dir)):
        if not fname.endswith(".pt"):
            continue
        fpath = f"models/{fname}"
        try:
            size_mb = round(_os.path.getsize(fpath) / (1024 * 1024), 1)
        except OSError:
            size_mb = 0.0
        result.append({
            "path": fpath,
            "filename": fname,
            "size_mb": size_mb,
            "is_active": fpath == active_path,
        })
    return {"models": result, "active_model": active_path}


@app.post("/api/models/activate")
async def activate_model(payload: dict, db: Session = Depends(get_db)):
    """Bascule le modèle actif sur le chemin fourni et l'applique à chaud."""
    import os as _os
    model_path = (payload.get("model_path") or "").strip()
    if not model_path:
        raise HTTPException(status_code=400, detail="model_path requis")
    if not _os.path.isfile(model_path):
        raise HTTPException(status_code=404, detail=f"Modèle introuvable : {model_path}")
    setting = db.query(models.SystemSetting).filter(models.SystemSetting.key == "detection_model_path").first()
    if setting:
        setting.value = model_path
    else:
        db.add(models.SystemSetting(key="detection_model_path", value=model_path))
    db.commit()
    v_engine = vision_engine.get_vision_engine()
    v_engine.apply_model_config(model_path=model_path)
    return {"activated": model_path, "message": f"Modèle basculé vers {model_path}"}


@app.post("/api/models/upload")
async def upload_model(file: UploadFile = File(...)):
    """Reçoit un fichier .pt et le sauvegarde dans le dossier models/."""
    import os as _os
    import shutil as _shutil
    if not file.filename or not file.filename.endswith(".pt"):
        raise HTTPException(status_code=400, detail="Seuls les fichiers .pt sont acceptés.")
    safe_name = _os.path.basename(file.filename)
    if not safe_name:
        raise HTTPException(status_code=400, detail="Nom de fichier invalide.")
    models_dir = "models"
    _os.makedirs(models_dir, exist_ok=True)
    dest = _os.path.join(models_dir, safe_name)
    with open(dest, "wb") as f:
        _shutil.copyfileobj(file.file, f)
    size_mb = round(_os.path.getsize(dest) / (1024 * 1024), 1)
    return {"path": f"models/{safe_name}", "filename": safe_name, "size_mb": size_mb}


@app.delete("/api/models/{filename}")
async def delete_model(filename: str):
    """Supprime un fichier .pt du dossier models/. Refuse si c'est le modèle actif."""
    import os as _os
    safe_name = _os.path.basename(filename)
    if not safe_name.endswith(".pt"):
        raise HTTPException(status_code=400, detail="Seuls les fichiers .pt peuvent être supprimés.")
    fpath = f"models/{safe_name}"
    v_engine = vision_engine.get_vision_engine()
    if fpath == v_engine.model_path:
        raise HTTPException(status_code=409, detail="Impossible de supprimer le modèle actuellement actif.")
    if not _os.path.isfile(fpath):
        raise HTTPException(status_code=404, detail=f"Modèle introuvable : {fpath}")
    _os.remove(fpath)
    return {"deleted": fpath}


# ─── IA Model Configuration ──────────────────────────────────────────────────
@app.get("/api/config/model", response_model=schemas.ModelConfig)
async def get_model_config(db: Session = Depends(get_db)):
    settings = db.query(models.SystemSetting).filter(
        models.SystemSetting.key.in_([
            "detection_model_path", "detection_threshold", "detection_nms_iou",
            "detection_max_det", "detection_imgsz", "tracking_persistence"
        ])
    ).all()
    config = {s.key: s.value for s in settings}

    return schemas.ModelConfig(
        selected_model=config.get("detection_model_path", "models/best_V5.pt"),
        confidence_threshold=float(config.get("detection_threshold", "0.7")),
        nms_iou_threshold=float(config.get("detection_nms_iou", "0.45")),
        max_detections=int(config.get("detection_max_det", "100")),
        inference_size=int(config.get("detection_imgsz", "1280")),
        tracking_persistence=config.get("tracking_persistence", "true").lower() == "true",
    )


@app.put("/api/config/model", response_model=schemas.ModelConfig)
async def update_model_config(config: schemas.ModelConfig, db: Session = Depends(get_db)):
    mapping = {
        "detection_model_path": config.selected_model,
        "detection_threshold": str(config.confidence_threshold),
        "detection_nms_iou": str(config.nms_iou_threshold),
        "detection_max_det": str(config.max_detections),
        "detection_imgsz": str(config.inference_size),
        "tracking_persistence": str(config.tracking_persistence).lower(),
    }
    for key, value in mapping.items():
        setting = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
        if setting:
            setting.value = value
        else:
            db.add(models.SystemSetting(key=key, value=value))
    db.commit()

    v_engine = vision_engine.get_vision_engine()
    v_engine.apply_model_config(
        model_path=config.selected_model,
        confidence_threshold=config.confidence_threshold,
        nms_iou_threshold=config.nms_iou_threshold,
        max_detections=config.max_detections,
        inference_size=config.inference_size,
        tracking_persistence=config.tracking_persistence,
    )
    return config


# ─── Virtual Line Configuration ──────────────────────────────────────────────
@app.get("/api/config/virtual-line", response_model=schemas.VirtualLineConfig)
async def get_virtual_line_config(db: Session = Depends(get_db)):
    settings = db.query(models.SystemSetting).filter(
        models.SystemSetting.key.in_([
            "virtual_line_y_percent", "virtual_line_span_percent", "virtual_line_direction"
        ])
    ).all()
    config = {s.key: s.value for s in settings}

    return schemas.VirtualLineConfig(
        position_percent=int(config.get("virtual_line_y_percent", "60")),
        line_span_percent=int(config.get("virtual_line_span_percent", "80")),
        direction=config.get("virtual_line_direction", "left-right"),
    )


@app.put("/api/config/virtual-line", response_model=schemas.VirtualLineConfig)
async def update_virtual_line_config(config: schemas.VirtualLineConfig, db: Session = Depends(get_db)):
    mapping = {
        "virtual_line_y_percent": str(config.position_percent),
        "virtual_line_span_percent": str(config.line_span_percent),
        "virtual_line_direction": config.direction,
    }
    for key, value in mapping.items():
        setting = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
        if setting:
            setting.value = value
        else:
            db.add(models.SystemSetting(key=key, value=value))
    db.commit()

    v_engine = vision_engine.get_vision_engine()
    v_engine.apply_virtual_line_config(
        position_percent=config.position_percent,
        line_span_percent=config.line_span_percent,
        direction=config.direction,
    )
    return config


@app.get("/api/config/line")
async def get_line_config(db: Session = Depends(get_db)):
    cfg = await get_virtual_line_config(db)
    line_type = "vertical" if cfg.direction in ["left-right", "right-left"] else "horizontal"
    return {
        "type": line_type,
        "direction": cfg.direction,
        "position_percent": cfg.position_percent,
        "line_span_percent": cfg.line_span_percent,
    }


@app.put("/api/config/line")
async def update_line_config(payload: dict, db: Session = Depends(get_db)):
    direction = payload.get("direction", "left-right")
    line_type = payload.get("type")
    position_percent = int(payload.get("position_percent", 60))
    line_span_percent = int(payload.get("line_span_percent", 80))

    # Force coherence between type and direction
    if direction in ["top-down", "bottom-up"]:
        line_type = "horizontal"
    elif direction in ["left-right", "right-left"]:
        line_type = "vertical"

    cfg = schemas.VirtualLineConfig(
        position_percent=position_percent,
        line_span_percent=line_span_percent,
        direction=direction,
    )
    await update_virtual_line_config(cfg, db)
    return {
        "type": line_type,
        "direction": direction,
        "position_percent": position_percent,
        "line_span_percent": line_span_percent,
    }


# ─── Auth ─────────────────────────────────────────────────────────────────────
@app.post("/token", response_model=schemas.Token)
async def login_for_access_token(request: Request, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    import datetime as _dt_auth
    user = db.query(models.User).filter(models.User.username == form_data.username).first()
    ip = request.client.host if request.client else None
    ua = request.headers.get("user-agent")
    if not user or not auth.verify_password(form_data.password, user.hashed_password):
        # Record failed login
        db.add(models.UserActivity(
            user_id=user.id if user else None,
            username=form_data.username,
            action="failed_login",
            ip_address=ip,
            user_agent=ua,
        ))
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    # Record successful login + update user stats
    user.last_login = _dt_auth.datetime.utcnow()
    user.login_count = (user.login_count or 0) + 1
    db.add(models.UserActivity(
        user_id=user.id,
        username=user.username,
        action="login",
        ip_address=ip,
        user_agent=ua,
    ))
    db.commit()
    access_token_expires = timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth.create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/users/me", response_model=schemas.User)
async def read_users_me(current_user: models.User = Depends(auth.get_current_user)):
    return current_user


@app.get("/api/users/me")
async def read_current_user_api(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    """Return current user profile + their role's permissions."""
    role = db.query(models.Role).filter(models.Role.name == current_user.role).first()
    permissions = json.loads(role.permissions) if role else []
    return {
        "id": current_user.id,
        "username": current_user.username,
        "full_name": current_user.full_name or current_user.username,
        "role": current_user.role,
        "is_active": current_user.is_active,
        "permissions": permissions,
    }


# ─── Sessions ─────────────────────────────────────────────────────────────────
@app.get("/sessions/active", response_model=schemas.Session | None)
async def read_active_session(db: Session = Depends(get_db)):
    return db.query(models.Session).filter(models.Session.status == "active").order_by(models.Session.start_time.desc()).first()


@app.get("/sessions/", response_model=schemas.SessionListResponse)
async def read_sessions(
    page: int = 1,
    page_size: int = 20,
    status: str | None = None,
    db: Session = Depends(get_db),
):
    page = max(page, 1)
    page_size = min(max(page_size, 1), 100)

    query = db.query(models.Session)
    if status:
        query = query.filter(models.Session.status == status)

    total = query.count()
    items = query.order_by(models.Session.start_time.desc()).offset((page - 1) * page_size).limit(page_size).all()
    active = db.query(models.Session).filter(models.Session.status == "active").first()
    return {
        "items": items,
        "total": total,
        "active_session_id": active.id if active else None,
    }


@app.post("/sessions/start", response_model=schemas.Session)
async def start_session(db: Session = Depends(get_db)):
    import datetime
    active = db.query(models.Session).filter(models.Session.status == "active").first()
    if active:
        return active

    session_id = f"S-{datetime.datetime.now().strftime('%Y%m%d-%H%M%S')}"
    db_session = models.Session(id=session_id)
    db.add(db_session)
    db.commit()
    db.refresh(db_session)

    v_engine = vision_engine.get_vision_engine()
    v_engine.set_active_session(session_id)

    # Notify all WebSocket clients so the UI resets immediately
    asyncio.run_coroutine_threadsafe(
        manager.broadcast({"type": "SESSION_STARTED", "data": {"session_id": session_id}}),
        _main_loop,
    ) if _main_loop else None

    return db_session


@app.post("/sessions/stop/{session_id}", response_model=schemas.Session)
async def stop_session(session_id: str, db: Session = Depends(get_db)):
    import datetime
    db_session = db.query(models.Session).filter(models.Session.id == session_id).first()
    if not db_session:
        raise HTTPException(status_code=404, detail="Session not found")
    if db_session.status != "active":
        return db_session

    db_session.end_time = datetime.datetime.utcnow()
    db_session.status = "completed"
    db.commit()
    db.refresh(db_session)

    v_engine = vision_engine.get_vision_engine()
    if v_engine.active_session_id == session_id:
        v_engine.set_active_session(None)

    # Notify all WebSocket clients with final counts so the UI can display them
    asyncio.run_coroutine_threadsafe(
        manager.broadcast({
            "type": "SESSION_STOPPED",
            "data": {
                "session_id": session_id,
                "total_count": db_session.total_count,
                "rejected_count": db_session.rejected_count,
            },
        }),
        _main_loop,
    ) if _main_loop else None

    return db_session


@app.delete("/api/sessions/{session_id}")
async def delete_session(session_id: str, db: Session = Depends(get_db)):
    db_session = db.query(models.Session).filter(models.Session.id == session_id).first()
    if not db_session:
        raise HTTPException(status_code=404, detail="Session not found")
    if db_session.status == "active":
        raise HTTPException(status_code=400, detail="Cannot delete active session")

    db.query(models.DetectionLog).filter(models.DetectionLog.session_id == session_id).delete()
    db.delete(db_session)
    db.commit()
    return {"deleted": 1, "session_id": session_id}


@app.delete("/api/sessions/{session_id}/logs")
async def clear_session_logs(session_id: str, db: Session = Depends(get_db)):
    """Delete all detection logs for a session without deleting the session itself."""
    db_session = db.query(models.Session).filter(models.Session.id == session_id).first()
    if not db_session:
        raise HTTPException(status_code=404, detail="Session not found")
    deleted = db.query(models.DetectionLog).filter(models.DetectionLog.session_id == session_id).delete()
    db_session.total_count = 0
    db_session.rejected_count = 0
    db.commit()
    return {"cleared": deleted, "session_id": session_id}


@app.delete("/api/sessions/batch")
async def delete_sessions_batch(payload: dict, db: Session = Depends(get_db)):
    session_ids = payload.get("session_ids", [])
    if not session_ids:
        return {"deleted": 0}

    active = db.query(models.Session).filter(models.Session.id.in_(session_ids), models.Session.status == "active").count()
    if active > 0:
        raise HTTPException(status_code=400, detail="Cannot delete active sessions")

    db.query(models.DetectionLog).filter(models.DetectionLog.session_id.in_(session_ids)).delete(synchronize_session=False)
    deleted = db.query(models.Session).filter(models.Session.id.in_(session_ids)).delete(synchronize_session=False)
    db.commit()
    return {"deleted": deleted}


# ─── Database Management ─────────────────────────────────────────────────────

_DB_PATH = "./cement_counter.db"


@app.get("/api/database/stats")
async def get_database_stats(db: Session = Depends(get_db)):
    """Real per-table stats + DB health (size, fragmentation, integrity)."""
    import os as _os, time as _time, psutil as _psutil
    from sqlalchemy import text
    from datetime import datetime as _dt

    # ── File size ─────────────────────────────────────────────────────────────
    db_size_bytes = _os.path.getsize(_DB_PATH) if _os.path.exists(_DB_PATH) else 0
    db_size_mb = round(db_size_bytes / (1024 * 1024), 2)

    # ── Fragmentation (SQLite PRAGMA) ─────────────────────────────────────────
    page_count  = db.execute(text("PRAGMA page_count")).scalar()  or 1
    freelist    = db.execute(text("PRAGMA freelist_count")).scalar() or 0
    page_size   = db.execute(text("PRAGMA page_size")).scalar()   or 4096
    frag_pct    = round(freelist / page_count * 100, 1)

    # ── Integrity check ───────────────────────────────────────────────────────
    integrity = db.execute(text("PRAGMA quick_check")).scalar() or "ok"

    # ── Per-table payload from dbstat virtual table ───────────────────────────
    try:
        dbstat_rows = db.execute(
            text("SELECT name, sum(payload) FROM dbstat WHERE aggregate=TRUE GROUP BY name")
        ).fetchall()
        tbl_bytes: dict[str, int] = {r[0]: int(r[1] or 0) for r in dbstat_rows}
    except Exception:
        tbl_bytes = {}

    # ── Row counts + last record timestamp ────────────────────────────────────
    table_defs = [
        ("detection_logs",  models.DetectionLog,  "timestamp"),
        ("sessions",        models.Session,        "start_time"),
        ("alert_history",   models.AlertHistory,   "timestamp"),
        ("alert_rules",     models.AlertRule,       None),
        ("quality_reviews", models.QualityReview,  "created_at"),
        ("system_settings", models.SystemSetting,  None),
        ("users",           models.User,            None),
    ]

    tables = []
    for tname, model, ts_col in table_defs:
        t0    = _time.perf_counter()
        count = db.query(func.count()).select_from(model).scalar() or 0
        q_ms  = round((_time.perf_counter() - t0) * 1000, 1)

        size_kb = round(tbl_bytes.get(tname, 0) / 1024, 1)

        last_ts = None
        if ts_col:
            try:
                col = getattr(model, ts_col)
                row = db.query(col).order_by(col.desc()).first()
                if row and row[0]:
                    last_ts = row[0].isoformat()
            except Exception:
                pass

        status = "fragmented" if frag_pct > 20 else ("attention" if frag_pct > 10 else "optimized")
        tables.append({
            "name":        tname,
            "rows":        count,
            "size_kb":     size_kb,
            "query_ms":    q_ms,
            "last_record": last_ts,
            "status":      status,
        })

    # ── Disk stats ────────────────────────────────────────────────────────────
    try:
        du = _psutil.disk_usage(".")
        disk_total_gb = round(du.total / (1024**3), 1)
        disk_used_gb  = round(du.used  / (1024**3), 1)
        disk_pct      = round(du.percent, 1)
    except Exception:
        disk_total_gb = disk_used_gb = disk_pct = 0

    # ── Retention setting ─────────────────────────────────────────────────────
    ret_setting = db.query(models.SystemSetting).filter_by(key="data_retention_days").first()
    retention_days = int(ret_setting.value) if ret_setting else 90

    return {
        "db_size_mb":      db_size_mb,
        "db_size_bytes":   db_size_bytes,
        "page_count":      page_count,
        "freelist_count":  freelist,
        "page_size":       page_size,
        "fragmentation_pct": frag_pct,
        "integrity":       integrity,
        "disk_total_gb":   disk_total_gb,
        "disk_used_gb":    disk_used_gb,
        "disk_pct":        disk_pct,
        "retention_days":  retention_days,
        "tables":          tables,
    }


@app.post("/api/database/optimize")
async def optimize_database():
    """Run VACUUM + ANALYZE to compact the DB and refresh query planner stats."""
    import os as _os, time as _time, sqlite3 as _sqlite3
    from sqlalchemy import text, create_engine as _ce

    size_before = _os.path.getsize(_DB_PATH) if _os.path.exists(_DB_PATH) else 0
    t0 = _time.perf_counter()

    # ANALYZE via SQLAlchemy
    _analyze_engine = _ce("sqlite:///" + _DB_PATH, connect_args={"check_same_thread": False})
    with _analyze_engine.connect() as conn:
        conn.execute(text("PRAGMA analysis_limit=400"))
        conn.execute(text("ANALYZE"))
        conn.commit()
    _analyze_engine.dispose()

    # VACUUM must run outside a transaction — use raw sqlite3
    _con = _sqlite3.connect(_DB_PATH, isolation_level=None)
    _con.execute("VACUUM")
    _con.close()

    elapsed_ms  = round((_time.perf_counter() - t0) * 1000)
    size_after  = _os.path.getsize(_DB_PATH) if _os.path.exists(_DB_PATH) else 0
    saved_kb    = round((size_before - size_after) / 1024, 1)

    return {
        "status":        "ok",
        "elapsed_ms":    elapsed_ms,
        "size_before_mb": round(size_before / (1024 * 1024), 2),
        "size_after_mb":  round(size_after  / (1024 * 1024), 2),
        "saved_kb":      max(saved_kb, 0),
    }


@app.post("/api/database/reindex")
async def reindex_database():
    """Rebuild all SQLite indexes."""
    import time as _time, sqlite3 as _sqlite3

    t0 = _time.perf_counter()
    _con = _sqlite3.connect(_DB_PATH, isolation_level=None)
    _con.execute("REINDEX")
    _con.close()
    elapsed_ms = round((_time.perf_counter() - t0) * 1000)

    return {"status": "ok", "elapsed_ms": elapsed_ms}


@app.get("/api/database/backup")
async def backup_database():
    """Stream the SQLite .db file as a direct download."""
    import os as _os
    from datetime import datetime as _dt
    from fastapi.responses import FileResponse

    if not _os.path.exists(_DB_PATH):
        raise HTTPException(status_code=404, detail="Fichier base de données introuvable")

    fname = f"cement_counter_backup_{_dt.utcnow().strftime('%Y%m%d_%H%M%S')}.db"
    return FileResponse(
        path=_DB_PATH,
        media_type="application/octet-stream",
        filename=fname,
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.post("/api/database/archive")
async def archive_old_sessions(payload: dict = None, db: Session = Depends(get_db)):
    """Delete completed sessions (and their logs) older than N days."""
    from datetime import datetime as _dt, timedelta as _td

    days   = int((payload or {}).get("days", 90))
    cutoff = _dt.utcnow() - _td(days=days)

    old_ids = [
        r[0] for r in
        db.query(models.Session.id)
          .filter(models.Session.status == "completed", models.Session.end_time <= cutoff)
          .all()
    ]
    if not old_ids:
        return {"archived_sessions": 0, "archived_logs": 0, "cutoff": cutoff.isoformat(), "days": days}

    deleted_logs = (
        db.query(models.DetectionLog)
          .filter(models.DetectionLog.session_id.in_(old_ids))
          .delete(synchronize_session=False)
    )
    deleted_sessions = (
        db.query(models.Session)
          .filter(models.Session.id.in_(old_ids))
          .delete(synchronize_session=False)
    )
    db.commit()

    return {
        "archived_sessions": deleted_sessions,
        "archived_logs":     deleted_logs,
        "cutoff":            cutoff.isoformat(),
        "days":              days,
    }


@app.post("/api/database/purge")
async def purge_old_logs(payload: dict = None, db: Session = Depends(get_db)):
    """Hard-delete detection logs (+ their quality reviews) older than N days."""
    import os as _os
    from datetime import datetime as _dt, timedelta as _td

    days   = int((payload or {}).get("days", 90))
    cutoff = _dt.utcnow() - _td(days=days)

    # Collect capture file paths before deletion
    urls = [
        r[0] for r in
        db.query(models.DetectionLog.capture_url)
          .filter(models.DetectionLog.timestamp <= cutoff,
                  models.DetectionLog.capture_url.isnot(None))
          .all()
    ]

    # Delete quality reviews linked to old logs (subquery)
    old_ids_subq = (
        db.query(models.DetectionLog.id)
          .filter(models.DetectionLog.timestamp <= cutoff)
          .subquery()
    )
    db.query(models.QualityReview)\
      .filter(models.QualityReview.log_id.in_(old_ids_subq))\
      .delete(synchronize_session=False)

    # Delete the logs themselves
    deleted = (
        db.query(models.DetectionLog)
          .filter(models.DetectionLog.timestamp <= cutoff)
          .delete(synchronize_session=False)
    )
    db.commit()

    # Try to remove orphaned capture files
    deleted_files = 0
    for url in urls:
        try:
            if url and _os.path.exists(url):
                _os.remove(url)
                deleted_files += 1
        except Exception:
            pass

    return {
        "purged_logs":  deleted,
        "purged_files": deleted_files,
        "cutoff":       cutoff.isoformat(),
        "days":         days,
    }


# ─── Diagnostics ─────────────────────────────────────────────────────────────

@app.get("/api/diagnostics/metrics")
async def get_diagnostic_metrics(db: Session = Depends(get_db)):
    """Real-time diagnostic KPIs: FPS, inference latency, precision, CPU/RAM."""
    import psutil as _psutil, time as _time
    from datetime import datetime as _dt, timedelta as _td

    v_eng        = vision_engine.get_vision_engine()
    engine_alive = bool(v_eng.running and v_eng.thread and v_eng.thread.is_alive())

    # ── FPS: try engine attribute, fallback to detection count in last 60s ────
    fps: float = 0.0
    try:
        fps = float(getattr(v_eng, "fps", 0) or 0)
    except Exception:
        fps = 0.0
    if fps == 0.0:
        since  = _dt.utcnow() - _td(seconds=60)
        count  = db.query(func.count(models.DetectionLog.id)).filter(
            models.DetectionLog.timestamp >= since
        ).scalar() or 0
        fps = round(count / 60, 1)

    # ── Inference ms: try engine attribute, fallback to avg interval proxy ────
    inference_ms = None
    try:
        raw = getattr(v_eng, "last_inference_ms", None)
        if raw:
            inference_ms = round(float(raw), 1)
    except Exception:
        pass
    if inference_ms is None:
        rows = (
            db.query(models.DetectionLog.interval)
            .filter(models.DetectionLog.interval.isnot(None),
                    models.DetectionLog.interval > 0)
            .order_by(models.DetectionLog.timestamp.desc())
            .limit(50).all()
        )
        if rows:
            avg_iv  = sum(r[0] for r in rows) / len(rows)
            fps_est = 1.0 / avg_iv if avg_iv > 0 else 0
            # Estimate: inference takes ~10% of inter-bag interval when running at full speed
            inference_ms = round(min(avg_iv * 1000 * 0.1, 999), 1)

    # ── Accuracy from all-time detection logs ─────────────────────────────────
    total     = db.query(func.count(models.DetectionLog.id)).scalar() or 0
    conformes = (
        db.query(func.count(models.DetectionLog.id))
        .filter(models.DetectionLog.status == "conforme").scalar() or 0
    )
    accuracy_pct = round(conformes / total * 100, 1) if total > 0 else None

    cpu = _psutil.cpu_percent(interval=None)
    mem = _psutil.virtual_memory()

    return {
        "fps":              fps,
        "inference_ms":     inference_ms,
        "accuracy_pct":     accuracy_pct,
        "cpu_pct":          round(cpu, 1),
        "ram_pct":          round(mem.percent, 1),
        "engine_alive":     engine_alive,
        "total_detections": total,
    }


@app.get("/api/diagnostics/logs")
async def get_diagnostic_logs(limit: int = 100, db: Session = Depends(get_db)):
    """Recent system events as a structured log stream."""
    events: list[dict] = []

    # AlertHistory → ERROR/WARN/INFO
    alerts = (
        db.query(models.AlertHistory)
        .order_by(models.AlertHistory.timestamp.desc())
        .limit(limit // 2).all()
    )
    for a in alerts:
        log_type = (
            "ERROR"   if a.alert_type == "critical" else
            "WARN"    if a.alert_type == "warning"  else
            "INFO"
        )
        events.append({
            "timestamp": a.timestamp.isoformat(),
            "time":      a.timestamp.strftime("%H:%M:%S"),
            "type":      log_type,
            "message":   f"[ALERTE] {a.title or 'Alerte système'} — {a.message}",
        })

    # Sessions → INFO (start) + SUCCESS (end)
    sessions = (
        db.query(models.Session)
        .order_by(models.Session.start_time.desc())
        .limit(limit // 4).all()
    )
    for s in sessions:
        total_bags = (s.total_count or 0) + (s.rejected_count or 0)
        events.append({
            "timestamp": s.start_time.isoformat(),
            "time":      s.start_time.strftime("%H:%M:%S"),
            "type":      "INFO",
            "message":   f"[SESSION] {s.id} démarrée — statut: {s.status}, {total_bags} sacs traités",
        })
        if s.end_time:
            events.append({
                "timestamp": s.end_time.isoformat(),
                "time":      s.end_time.strftime("%H:%M:%S"),
                "type":      "SUCCESS",
                "message":   f"[SESSION] {s.id} terminée — {s.total_count} conformes, {s.rejected_count} rejetés",
            })

    events.sort(key=lambda e: e["timestamp"], reverse=True)
    return {"logs": events[:limit], "total": len(events)}


@app.get("/api/diagnostics/logs/download")
async def download_diagnostic_logs(db: Session = Depends(get_db)):
    """Download system logs as a .txt file."""
    from datetime import datetime as _dt
    result = await get_diagnostic_logs(limit=500, db=db)
    lines  = [f"=== Logs Diagnostics — {_dt.utcnow().strftime('%Y-%m-%d %H:%M UTC')} ==="]
    for e in result["logs"]:
        lines.append(f"[{e['timestamp']}] {e['type']:<8} {e['message']}")
    fname = f"diagnostics_{_dt.utcnow().strftime('%Y%m%d_%H%M%S')}.txt"
    return Response(
        content="\n".join(lines).encode("utf-8"),
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.post("/api/diagnostics/run-tests")
async def run_diagnostic_tests(payload: dict = None, db: Session = Depends(get_db)):
    """Run real component tests. Optional body: {"test": "yolo|db|disk|api|camera"}."""
    import time as _time, os as _os, tempfile as _tmp
    from datetime import datetime as _dt

    only = (payload or {}).get("test")  # if set, run only this test key

    v_eng        = vision_engine.get_vision_engine()
    engine_alive = bool(v_eng.running and v_eng.thread and v_eng.thread.is_alive())

    def _run(key: str) -> dict:
        if key == "yolo":
            t0     = _time.perf_counter()
            alive  = engine_alive
            ms     = round((_time.perf_counter() - t0) * 1000, 1)
            model  = _os.path.basename(getattr(v_eng, "model_path", "") or "—")
            return {
                "name":   "Moteur IA (YOLOv11)",
                "key":    "yolo",
                "status": "pass" if alive else "fail",
                "metric": model,
                "latency": f"{ms} ms",
                "detail": "En ligne" if alive else "Hors ligne — démarrez une session",
            }

        if key == "db":
            t0 = _time.perf_counter()
            try:
                test_key = "_diag_write_test_"
                ex = db.query(models.SystemSetting).filter_by(key=test_key).first()
                if ex:
                    ex.value = _dt.utcnow().isoformat()
                else:
                    db.add(models.SystemSetting(key=test_key, value=_dt.utcnow().isoformat()))
                db.commit()
                db.query(models.SystemSetting).filter_by(key=test_key).delete()
                db.commit()
                ok, detail = True, "Lecture/écriture OK"
            except Exception as e:
                ok, detail = False, str(e)
            ms = round((_time.perf_counter() - t0) * 1000, 1)
            return {
                "name":    "Base de Données (SQLite)",
                "key":     "db",
                "status":  "pass" if ok else "fail",
                "metric":  f"{ms} ms",
                "latency": f"{ms} ms",
                "detail":  detail,
            }

        if key == "disk":
            t0  = _time.perf_counter()
            buf = b"\x00" * (1024 * 1024)
            try:
                with _tmp.NamedTemporaryFile(delete=False, suffix=".tmp") as f:
                    fname = f.name; f.write(buf)
                elapsed = _time.perf_counter() - t0
                _os.remove(fname)
                speed = round(1.0 / elapsed, 1)
                ok, detail = True, f"{speed} MB/s"
            except Exception as e:
                speed, ok, detail = 0, False, str(e)
            ms = round((_time.perf_counter() - t0) * 1000, 1)
            return {
                "name":    "Écriture Disque",
                "key":     "disk",
                "status":  "pass" if ok else "fail",
                "metric":  f"{speed} MB/s",
                "latency": f"{speed} MB/s",
                "detail":  detail,
            }

        if key == "api":
            t0 = _time.perf_counter()
            try:
                db.query(func.count(models.DetectionLog.id)).scalar()
                ok, detail = True, "Endpoint opérationnel"
            except Exception as e:
                ok, detail = False, str(e)
            ms = round((_time.perf_counter() - t0) * 1000, 1)
            return {
                "name":    "API FastAPI",
                "key":     "api",
                "status":  "pass" if ok else "fail",
                "metric":  f"{ms} ms",
                "latency": f"{ms} ms",
                "detail":  detail,
            }

        if key == "camera":
            # Read saved camera config from SystemSettings
            cam_cfg = {
                s.key: s.value for s in
                db.query(models.SystemSetting).filter(
                    models.SystemSetting.key.in_(["camera_source_type", "camera_url"])
                ).all()
            }
            src_type = cam_cfg.get("camera_source_type", "webcam")
            cam_url  = cam_cfg.get("camera_url", "0")

            if not cam_url:
                return {
                    "name":    "Flux Vidéo (Caméra)",
                    "key":     "camera",
                    "status":  "warn",
                    "metric":  "—",
                    "latency": "—",
                    "detail":  "Aucune source caméra configurée",
                }

            # Build a minimal CameraConfig-like object and call _test_camera_sync
            import importlib as _imp, time as _time2
            cam_config_obj = type("CC", (), {"source_type": src_type, "url": cam_url})()
            t0 = _time2.perf_counter()
            try:
                cam_result = _test_camera_sync(cam_config_obj)
            except Exception as e:
                cam_result = {"success": False, "message": str(e)}
            ms = round((_time2.perf_counter() - t0) * 1000)

            ok     = cam_result.get("success", False)
            res    = cam_result.get("resolution_detected", "—")
            fps_v  = cam_result.get("fps_detected", "—")
            msg    = cam_result.get("message", "—")
            metric = f"{res} @ {fps_v} FPS" if ok else "Échec"
            return {
                "name":    "Flux Vidéo (Caméra)",
                "key":     "camera",
                "status":  "pass" if ok else "fail",
                "metric":  metric,
                "latency": f"{ms} ms",
                "detail":  msg,
            }

        return {"key": key, "status": "unknown"}

    all_keys = ["yolo", "db", "disk", "api", "camera"]
    results  = [_run(only)] if only and only in all_keys else [_run(k) for k in all_keys]
    passed   = sum(1 for r in results if r["status"] == "pass")
    return {
        "results":  results,
        "passed":   passed,
        "total":    len(results),
        "all_pass": passed == len(results),
    }


@app.post("/api/diagnostics/benchmark")
async def run_ia_benchmark():
    """Quick IA benchmark: 20 inferences on a blank frame."""
    import time as _time, os as _os
    import numpy as _np

    v_eng = vision_engine.get_vision_engine()
    if not (getattr(v_eng, "running", False) and getattr(v_eng, "model_path", None)):
        raise HTTPException(
            status_code=503,
            detail="Moteur IA non actif — démarrez une session d'abord",
        )

    n_frames = 20
    try:
        from ultralytics import YOLO as _YOLO
        model        = _YOLO(v_eng.model_path)
        blank_frame  = _np.zeros((640, 640, 3), dtype=_np.uint8)
        times_ms: list[float] = []
        for _ in range(n_frames):
            t0 = _time.perf_counter()
            model.predict(blank_frame, verbose=False, imgsz=640)
            times_ms.append((_time.perf_counter() - t0) * 1000)
        del model

        avg_ms  = round(sum(times_ms) / len(times_ms), 1)
        min_ms  = round(min(times_ms), 1)
        max_ms  = round(max(times_ms), 1)
        fps_eq  = round(1000 / avg_ms, 1) if avg_ms > 0 else 0
        return {
            "status":        "ok",
            "frames_tested": n_frames,
            "avg_ms":        avg_ms,
            "min_ms":        min_ms,
            "max_ms":        max_ms,
            "fps_equiv":     fps_eq,
            "model":         _os.path.basename(v_eng.model_path),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Benchmark échoué : {e}")


# ─── Template & Colour Configuration ─────────────────────────────────────────
import json as _json_tmpl_ep
import colorsys as _colorsys
import os as _os_tmpl

_TEMPLATES_DIR = "backend/static/templates"
_os_tmpl.makedirs(_TEMPLATES_DIR, exist_ok=True)


def _hex_to_hsv_range(hex_color: str, tolerance: int) -> dict:
    """Convert a hex colour + tolerance into an OpenCV HSV range dict."""
    hex_color = hex_color.lstrip("#")
    r, g, b = (int(hex_color[i:i+2], 16) / 255.0 for i in (0, 2, 4))
    h, s, v = _colorsys.rgb_to_hsv(r, g, b)
    h_cv = h * 180.0          # OpenCV H: 0-180
    s_cv = s * 255.0           # OpenCV S: 0-255
    v_cv = v * 255.0           # OpenCV V: 0-255
    h_tol  = tolerance * 0.9    # H is compact: small tolerance
    sv_tol = tolerance * 2.0    # S and V have larger tolerance
    return {
        "h_min": max(0,   int(h_cv - h_tol)),
        "h_max": min(180, int(h_cv + h_tol)),
        "s_min": max(0,   int(s_cv - sv_tol)),
        "s_max": min(255, int(s_cv + sv_tol)),
        "v_min": max(0,   int(v_cv - sv_tol)),
        "v_max": min(255, int(v_cv + sv_tol)),
    }


def _get_setting(db, key: str, default: str = "") -> str:
    s = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
    return s.value if s else default


def _set_setting(db, key: str, value: str):
    s = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
    if s:
        s.value = value
    else:
        db.add(models.SystemSetting(key=key, value=value))


@app.get("/api/config/template")
async def get_template_config(db: Session = Depends(get_db)):
    active_file = _get_setting(db, "template_active_file")
    threshold   = float(_get_setting(db, "template_threshold", "0.65"))
    color_thr   = float(_get_setting(db, "template_color_threshold", "0.25"))
    color_refs  = _json_tmpl_ep.loads(_get_setting(db, "template_colors", "[]"))

    # Build history from files on disk
    history = []
    try:
        for fname in sorted(_os_tmpl.listdir(_TEMPLATES_DIR), reverse=True):
            if fname.lower().endswith((".jpg", ".jpeg", ".png")):
                fpath = _os_tmpl.path.join(_TEMPLATES_DIR, fname)
                img = __import__("cv2").imread(fpath)
                w, h = (img.shape[1], img.shape[0]) if img is not None else (0, 0)
                history.append({
                    "filename": fname,
                    "url": f"/static/templates/{fname}",
                    "size_kb": round(_os_tmpl.path.getsize(fpath) / 1024, 1),
                    "width": w,
                    "height": h,
                    "is_active": fname == active_file,
                })
    except Exception:
        pass

    active_meta = next((t for t in history if t["is_active"]), None)
    return {
        "active_file": active_file,
        "active_url": f"/static/templates/{active_file}" if active_file else None,
        "active_width":  active_meta["width"]  if active_meta else 0,
        "active_height": active_meta["height"] if active_meta else 0,
        "threshold": threshold,
        "color_threshold": color_thr,
        "color_refs": color_refs,
        "history": history,
    }


@app.post("/api/config/template/upload")
async def upload_template(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Seules les images sont acceptées (JPEG, PNG).")
    import datetime as _dt_tmpl
    ext      = _os_tmpl.path.splitext(file.filename or "img.jpg")[1].lower() or ".jpg"
    fname    = f"template_{_dt_tmpl.datetime.utcnow().strftime('%Y%m%d_%H%M%S')}{ext}"
    fpath    = _os_tmpl.path.join(_TEMPLATES_DIR, fname)
    contents = await file.read()
    with open(fpath, "wb") as f:
        f.write(contents)

    # Read back to verify and get dimensions
    import cv2 as _cv2_t
    img = _cv2_t.imread(fpath)
    if img is None:
        _os_tmpl.remove(fpath)
        raise HTTPException(status_code=400, detail="Fichier image invalide ou corrompu.")

    # Set as active
    _set_setting(db, "template_active_file", fname)
    threshold  = float(_get_setting(db, "template_threshold", "0.65"))
    color_thr  = float(_get_setting(db, "template_color_threshold", "0.25"))
    color_refs = _json_tmpl_ep.loads(_get_setting(db, "template_colors", "[]"))
    db.commit()

    # Apply to running engine
    vision_engine.get_vision_engine().apply_template_config(fpath, threshold, color_refs, color_thr)

    return {
        "filename": fname,
        "url": f"/static/templates/{fname}",
        "width": img.shape[1],
        "height": img.shape[0],
    }


@app.post("/api/config/template/activate/{filename}")
async def activate_template(filename: str, db: Session = Depends(get_db)):
    fpath = _os_tmpl.path.join(_TEMPLATES_DIR, filename)
    if not _os_tmpl.path.exists(fpath):
        raise HTTPException(status_code=404, detail="Template introuvable.")
    _set_setting(db, "template_active_file", filename)
    threshold  = float(_get_setting(db, "template_threshold", "0.65"))
    color_thr  = float(_get_setting(db, "template_color_threshold", "0.25"))
    color_refs = _json_tmpl_ep.loads(_get_setting(db, "template_colors", "[]"))
    db.commit()
    vision_engine.get_vision_engine().apply_template_config(fpath, threshold, color_refs, color_thr)
    return {"ok": True, "active_file": filename}


@app.delete("/api/config/template/history/{filename}")
async def delete_template(filename: str, db: Session = Depends(get_db)):
    fpath = _os_tmpl.path.join(_TEMPLATES_DIR, filename)
    if not _os_tmpl.path.exists(fpath):
        raise HTTPException(status_code=404, detail="Template introuvable.")
    active = _get_setting(db, "template_active_file")
    if filename == active:
        raise HTTPException(status_code=400, detail="Impossible de supprimer le template actif. Activez un autre d'abord.")
    _os_tmpl.remove(fpath)
    return {"ok": True}


@app.put("/api/config/template/settings")
async def update_template_settings(body: dict, db: Session = Depends(get_db)):
    """Update threshold and/or colour score threshold."""
    if "threshold" in body:
        _set_setting(db, "template_threshold", str(float(body["threshold"])))
    if "color_threshold" in body:
        _set_setting(db, "template_color_threshold", str(float(body["color_threshold"])))
    db.commit()
    # Re-apply to engine with current template
    active_file = _get_setting(db, "template_active_file")
    fpath       = _os_tmpl.path.join(_TEMPLATES_DIR, active_file) if active_file else None
    threshold   = float(_get_setting(db, "template_threshold", "0.65"))
    color_thr   = float(_get_setting(db, "template_color_threshold", "0.25"))
    color_refs  = _json_tmpl_ep.loads(_get_setting(db, "template_colors", "[]"))
    vision_engine.get_vision_engine().apply_template_config(fpath, threshold, color_refs, color_thr)
    return {"ok": True, "threshold": threshold, "color_threshold": color_thr}


@app.get("/api/config/colors")
async def get_color_refs(db: Session = Depends(get_db)):
    refs = _json_tmpl_ep.loads(_get_setting(db, "template_colors", "[]"))
    return refs


@app.post("/api/config/colors")
async def add_color_ref(body: dict, db: Session = Depends(get_db)):
    """Body: {name, hex, tolerance (0-100)}"""
    name      = body.get("name", "Couleur").strip()
    hex_color = body.get("hex", "#808080")
    tolerance = int(body.get("tolerance", 25))
    hsv_range = _hex_to_hsv_range(hex_color, tolerance)
    refs = _json_tmpl_ep.loads(_get_setting(db, "template_colors", "[]"))
    refs.append({"name": name, "hex": hex_color, "tolerance": tolerance, **hsv_range})
    _set_setting(db, "template_colors", _json_tmpl_ep.dumps(refs))
    db.commit()
    # Re-apply
    active_file = _get_setting(db, "template_active_file")
    fpath       = _os_tmpl.path.join(_TEMPLATES_DIR, active_file) if active_file else None
    threshold   = float(_get_setting(db, "template_threshold", "0.65"))
    color_thr   = float(_get_setting(db, "template_color_threshold", "0.25"))
    vision_engine.get_vision_engine().apply_template_config(fpath, threshold, refs, color_thr)
    return refs


@app.put("/api/config/colors/{idx}")
async def update_color_ref(idx: int, body: dict, db: Session = Depends(get_db)):
    refs = _json_tmpl_ep.loads(_get_setting(db, "template_colors", "[]"))
    if idx < 0 or idx >= len(refs):
        raise HTTPException(status_code=404, detail="Référence couleur introuvable.")
    if "name"      in body: refs[idx]["name"]      = body["name"]
    if "tolerance" in body:
        tol = int(body["tolerance"])
        refs[idx]["tolerance"] = tol
        refs[idx].update(_hex_to_hsv_range(refs[idx]["hex"], tol))
    if "hex" in body:
        refs[idx]["hex"] = body["hex"]
        refs[idx].update(_hex_to_hsv_range(body["hex"], refs[idx].get("tolerance", 25)))
    _set_setting(db, "template_colors", _json_tmpl_ep.dumps(refs))
    db.commit()
    active_file = _get_setting(db, "template_active_file")
    fpath       = _os_tmpl.path.join(_TEMPLATES_DIR, active_file) if active_file else None
    threshold   = float(_get_setting(db, "template_threshold", "0.65"))
    color_thr   = float(_get_setting(db, "template_color_threshold", "0.25"))
    vision_engine.get_vision_engine().apply_template_config(fpath, threshold, refs, color_thr)
    return refs


@app.delete("/api/config/colors/{idx}")
async def delete_color_ref(idx: int, db: Session = Depends(get_db)):
    refs = _json_tmpl_ep.loads(_get_setting(db, "template_colors", "[]"))
    if idx < 0 or idx >= len(refs):
        raise HTTPException(status_code=404, detail="Référence couleur introuvable.")
    refs.pop(idx)
    _set_setting(db, "template_colors", _json_tmpl_ep.dumps(refs))
    db.commit()
    active_file = _get_setting(db, "template_active_file")
    fpath       = _os_tmpl.path.join(_TEMPLATES_DIR, active_file) if active_file else None
    threshold   = float(_get_setting(db, "template_threshold", "0.65"))
    color_thr   = float(_get_setting(db, "template_color_threshold", "0.25"))
    vision_engine.get_vision_engine().apply_template_config(fpath, threshold, refs, color_thr)
    return refs


# ─── Device Management ────────────────────────────────────────────────────────
import datetime as _dt_dev

def _cam_to_dict(c: models.Camera) -> dict:
    return {
        "id": c.id,
        "name": c.name,
        "source_type": c.source_type,
        "url": c.url,
        "resolution": c.resolution,
        "fps": c.fps,
        "is_active": c.is_active,
        "notes": c.notes,
        "created_at": c.created_at.isoformat() if c.created_at else None,
        "last_tested_at": c.last_tested_at.isoformat() if c.last_tested_at else None,
        "last_status": c.last_status,
        "last_latency_ms": c.last_latency_ms,
    }


@app.get("/api/devices/cameras")
async def list_cameras(db: Session = Depends(get_db)):
    cameras = db.query(models.Camera).order_by(models.Camera.id).all()
    return [_cam_to_dict(c) for c in cameras]


@app.post("/api/devices/cameras", status_code=201)
async def create_camera(payload: schemas.CameraDeviceCreate, db: Session = Depends(get_db)):
    cam = models.Camera(
        name=payload.name,
        source_type=payload.source_type,
        url=payload.url,
        resolution=payload.resolution,
        fps=payload.fps,
        notes=payload.notes,
        is_active=False,
    )
    db.add(cam)
    db.commit()
    db.refresh(cam)
    return _cam_to_dict(cam)


@app.put("/api/devices/cameras/{cam_id}")
async def update_camera(cam_id: int, payload: schemas.CameraDeviceUpdate, db: Session = Depends(get_db)):
    cam = db.query(models.Camera).filter(models.Camera.id == cam_id).first()
    if not cam:
        raise HTTPException(status_code=404, detail="Caméra introuvable")
    for field, val in payload.model_dump(exclude_unset=True).items():
        setattr(cam, field, val)
    db.commit()
    db.refresh(cam)
    return _cam_to_dict(cam)


@app.delete("/api/devices/cameras/{cam_id}")
async def delete_camera(cam_id: int, db: Session = Depends(get_db)):
    cam = db.query(models.Camera).filter(models.Camera.id == cam_id).first()
    if not cam:
        raise HTTPException(status_code=404, detail="Caméra introuvable")
    if cam.is_active:
        raise HTTPException(status_code=400, detail="Impossible de supprimer la caméra active. Activez une autre caméra d'abord.")
    db.delete(cam)
    db.commit()
    return {"ok": True}


@app.post("/api/devices/cameras/{cam_id}/test")
async def test_camera_device(cam_id: int, db: Session = Depends(get_db)):
    cam = db.query(models.Camera).filter(models.Camera.id == cam_id).first()
    if not cam:
        raise HTTPException(status_code=404, detail="Caméra introuvable")
    config_obj = type("CC", (), {"source_type": cam.source_type, "url": cam.url})()
    loop = asyncio.get_event_loop()
    import time as _t_dev
    t0 = _t_dev.perf_counter()
    result = await loop.run_in_executor(None, _test_camera_sync, config_obj)
    latency_ms = round((_t_dev.perf_counter() - t0) * 1000)
    cam.last_tested_at = _dt_dev.datetime.utcnow()
    cam.last_status = "online" if result.get("success") else "offline"
    cam.last_latency_ms = latency_ms if result.get("success") else None
    db.commit()
    return {**result, "camera_id": cam_id, "latency_ms": latency_ms}


@app.post("/api/devices/cameras/{cam_id}/activate")
async def activate_camera(cam_id: int, db: Session = Depends(get_db)):
    cam = db.query(models.Camera).filter(models.Camera.id == cam_id).first()
    if not cam:
        raise HTTPException(status_code=404, detail="Caméra introuvable")
    # Deactivate all, then activate this one
    db.query(models.Camera).update({"is_active": False})
    cam.is_active = True
    db.commit()
    # Sync legacy SystemSettings keys for backward compat
    _mapping = {
        "camera_source_type": cam.source_type,
        "camera_url": cam.url,
        "camera_resolution": cam.resolution,
        "camera_fps": str(cam.fps),
    }
    for key, value in _mapping.items():
        s = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
        if s:
            s.value = value
        else:
            db.add(models.SystemSetting(key=key, value=value))
    db.commit()
    # Apply to running vision engine
    v_engine = vision_engine.get_vision_engine()
    new_source = int(cam.url) if cam.source_type == "webcam" and cam.url.isdigit() else cam.url
    source_changed = v_engine.video_source != new_source
    if source_changed:
        v_engine.video_source = new_source
    cam_settings = db.query(models.SystemSetting).filter(
        models.SystemSetting.key.in_(["camera_brightness", "camera_contrast", "camera_autofocus"])
    ).all()
    cs = {s.key: s.value for s in cam_settings}
    v_engine.apply_camera_settings(
        resolution=cam.resolution,
        fps=cam.fps,
        brightness=int(cs.get("camera_brightness", "50")),
        contrast=int(cs.get("camera_contrast", "65")),
        autofocus=cs.get("camera_autofocus", "true").lower() == "true",
    )
    if v_engine.running and source_changed:
        if v_engine.stop():
            v_engine.start()
    return _cam_to_dict(cam)


@app.get("/api/devices/system")
async def get_device_system():
    import psutil as _ps_dev
    cpu_pct = _ps_dev.cpu_percent(interval=0.2)
    mem = _ps_dev.virtual_memory()
    disk = _ps_dev.disk_usage(".")
    temp_c = None
    try:
        temps = _ps_dev.sensors_temperatures()
        if temps:
            for key in ("coretemp", "cpu_thermal", "k10temp"):
                if key in temps and temps[key]:
                    temp_c = round(temps[key][0].current, 1)
                    break
    except Exception:
        pass
    return {
        "cpu_pct": round(cpu_pct, 1),
        "cpu_temp_c": temp_c,
        "ram_used_gb": round(mem.used / 1024 ** 3, 1),
        "ram_total_gb": round(mem.total / 1024 ** 3, 1),
        "ram_pct": round(mem.percent, 1),
        "disk_used_gb": round(disk.used / 1024 ** 3, 1),
        "disk_total_gb": round(disk.total / 1024 ** 3, 1),
        "disk_pct": round(disk.percent, 1),
    }


@app.get("/api/devices/services")
async def get_device_services():
    import sqlite3 as _sq3_dev
    v_engine = vision_engine.get_vision_engine()
    yolo_alive = False
    try:
        yolo_alive = v_engine.running and v_engine.thread is not None and v_engine.thread.is_alive()
    except Exception:
        pass
    db_ok = False
    try:
        _c = _sq3_dev.connect(_DB_PATH, timeout=2)
        _c.execute("SELECT 1")
        _c.close()
        db_ok = True
    except Exception:
        pass
    return [
        {"name": "Inférence YOLO", "key": "yolo",    "status": "running" if yolo_alive else "stopped"},
        {"name": "Base de Données SQLite", "key": "sqlite",  "status": "running" if db_ok else "error"},
        {"name": "API FastAPI",            "key": "fastapi", "status": "running"},
        {"name": "Flux WebSocket/MJPEG",   "key": "mjpeg",   "status": "running" if v_engine.running else "warning"},
    ]


# ─── Entrypoint ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        app,
        host="0.0.0.0",
        port=8000,
        timeout_graceful_shutdown=5,
    )
